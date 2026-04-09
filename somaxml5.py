# Analisador Fiscal Profissional — NF-e / CT-e / NFSe / MDF-e
# ============================================================

import tkinter as tk
from tkinter import filedialog, messagebox
import tkinter.ttk as ttk
import xml.etree.ElementTree as ET
import zipfile
import os
import csv
import json
import logging
import threading
import queue
from collections import defaultdict
from datetime import datetime
from pathlib import Path

# Importações opcionais com fallback
try:
    import rarfile
    RARFILE_AVAILABLE = True
except ImportError:
    RARFILE_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference, PieChart
    from openpyxl.chart.series import DataPoint
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, HRFlowable, KeepTogether
    )
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# ─── Config Logging ────────────────────────────────────────────────────────────
LOG_FILENAME = "erros_fiscal.log"
logging.basicConfig(
    filename=LOG_FILENAME,
    level=logging.INFO,
    format="%(asctime)s %(levelname)s: %(message)s",
    encoding="utf-8",
)

# ─── Paleta de Cores Profissional ───────────────────────────────────────────────
COLORS = {
    # Fundos
    "bg_main":       "#0F1117",
    "bg_sidebar":    "#13151E",
    "bg_card":       "#1A1D2E",
    "bg_card_hover": "#1E2235",
    "bg_input":      "#252840",
    "bg_table":      "#161927",
    "bg_table_alt":  "#1A1D2E",
    "bg_header":     "#0D0F1A",

    # Acentos
    "accent":        "#4F7EFF",
    "accent_hover":  "#6B94FF",
    "accent_dim":    "#2A4499",
    "accent2":       "#00D4AA",   # verde-água para autorizado
    "accent3":       "#FF6B6B",   # vermelho para erro/cancelado
    "accent4":       "#FFB347",   # laranja para alerta
    "accent5":       "#A78BFA",   # roxo para NFSe

    # Texto
    "text_primary":  "#E8EAF6",
    "text_secondary":"#8B90C4",
    "text_dim":      "#4A4F7A",
    "text_white":    "#FFFFFF",

    # Bordas
    "border":        "#252840",
    "border_accent": "#4F7EFF",

    # Status
    "status_ok":     "#00D4AA",
    "status_warn":   "#FFB347",
    "status_error":  "#FF6B6B",
    "status_cancel": "#8B90C4",
}

# ─── Fontes ─────────────────────────────────────────────────────────────────────
FONT_TITLE   = ("Segoe UI", 18, "bold")
FONT_HEADING = ("Segoe UI", 12, "bold")
FONT_BODY    = ("Segoe UI", 10)
FONT_SMALL   = ("Segoe UI", 9)
FONT_MONO    = ("Consolas", 9)
FONT_LABEL   = ("Segoe UI", 10, "bold")
FONT_KPI     = ("Segoe UI", 22, "bold")
FONT_KPI_SUB = ("Segoe UI", 10)

# ─── Global ─────────────────────────────────────────────────────────────────────
GLOBAL_ITEM_DETAILS = defaultdict(list)
APP_VERSION = "7.1.0"
CONFIG_FILE = "fiscal_config.json"


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  Widgets Customizados                                                        ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

class StyledButton(tk.Frame):
    """Botão estilizado com hover e ícone opcional."""
    def __init__(self, parent, text="", icon="", command=None,
                 style="primary", width=None, **kwargs):
        super().__init__(parent, bg=COLORS["bg_main"], **kwargs)
        self.command = command

        bg_map = {
            "primary":   (COLORS["accent"],       COLORS["accent_hover"]),
            "secondary": (COLORS["bg_card"],       COLORS["bg_card_hover"]),
            "danger":    (COLORS["accent3"],       "#FF8F8F"),
            "success":   (COLORS["accent2"],       "#00ECC0"),
            "warning":   (COLORS["accent4"],       "#FFC970"),
        }
        self._bg, self._bg_h = bg_map.get(style, bg_map["primary"])
        self._fg = COLORS["text_white"] if style != "secondary" else COLORS["text_primary"]

        label_text = f"{icon}  {text}" if icon else text
        self._lbl = tk.Label(
            self, text=label_text,
            bg=self._bg, fg=self._fg,
            font=FONT_BODY, padx=16, pady=8,
            cursor="hand2"
        )
        if width:
            self._lbl.configure(width=width)
        self._lbl.pack(fill="both", expand=True)

        self._lbl.bind("<Enter>",    self._on_enter)
        self._lbl.bind("<Leave>",    self._on_leave)
        self._lbl.bind("<Button-1>", self._on_click)

    def _on_enter(self, e):
        self._lbl.configure(bg=self._bg_h)
    def _on_leave(self, e):
        self._lbl.configure(bg=self._bg)
    def _on_click(self, e):
        if self.command:
            self.command()

    def configure_state(self, enabled=True):
        if enabled:
            self._lbl.configure(bg=self._bg, cursor="hand2")
        else:
            self._lbl.configure(bg=COLORS["text_dim"], cursor="")


class KPICard(tk.Frame):
    """Card de indicador KPI."""
    def __init__(self, parent, title="", value="—", unit="", color=None, **kwargs):
        c = color or COLORS["accent"]
        super().__init__(parent, bg=COLORS["bg_card"],
                         highlightbackground=c, highlightthickness=1, **kwargs)

        # Barra lateral colorida
        bar = tk.Frame(self, bg=c, width=4)
        bar.pack(side="left", fill="y")

        inner = tk.Frame(self, bg=COLORS["bg_card"], padx=16, pady=12)
        inner.pack(side="left", fill="both", expand=True)

        self._val_var = tk.StringVar(value=value)
        self._sub_var = tk.StringVar(value=title)

        tk.Label(inner, textvariable=self._val_var,
                 bg=COLORS["bg_card"], fg=c,
                 font=FONT_KPI).pack(anchor="w")
        tk.Label(inner, textvariable=self._sub_var,
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"],
                 font=FONT_KPI_SUB).pack(anchor="w")

    def update(self, value, subtitle=None):
        self._val_var.set(value)
        if subtitle:
            self._sub_var.set(subtitle)


class StatusBar(tk.Frame):
    """Barra de status inferior."""
    def __init__(self, parent, **kwargs):
        super().__init__(parent, bg=COLORS["bg_header"],
                         height=28, **kwargs)
        self.pack_propagate(False)
        self._msg = tk.StringVar(value="Pronto")
        self._progress_var = tk.DoubleVar(value=0)

        tk.Label(self, textvariable=self._msg,
                 bg=COLORS["bg_header"], fg=COLORS["text_secondary"],
                 font=FONT_SMALL, padx=12).pack(side="left")

        self._progress = ttk.Progressbar(
            self, variable=self._progress_var,
            maximum=100, length=200, mode="determinate"
        )
        self._progress.pack(side="right", padx=12, pady=4)

    def set_message(self, msg):
        self._msg.set(msg)

    def set_progress(self, val):
        self._progress_var.set(val)

    def reset(self):
        self._msg.set("Pronto")
        self._progress_var.set(0)


class FilterBar(tk.Frame):
    """Barra de filtros para a tabela."""
    def __init__(self, parent, on_filter=None, **kwargs):
        super().__init__(parent, bg=COLORS["bg_card"], **kwargs)
        self.on_filter = on_filter

        # Busca
        tk.Label(self, text="🔍", bg=COLORS["bg_card"],
                 fg=COLORS["text_secondary"], font=FONT_BODY).pack(side="left", padx=(10,2))
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self._trigger)
        search_entry = tk.Entry(self, textvariable=self.search_var,
                                bg=COLORS["bg_input"], fg=COLORS["text_primary"],
                                insertbackground=COLORS["text_primary"],
                                relief="flat", font=FONT_BODY, width=24)
        search_entry.pack(side="left", padx=4, pady=6)

        # Tipo
        tk.Label(self, text="Tipo:", bg=COLORS["bg_card"],
                 fg=COLORS["text_secondary"], font=FONT_SMALL).pack(side="left", padx=(12,2))
        self.tipo_var = tk.StringVar(value="Todos")
        tipo_cb = ttk.Combobox(self, textvariable=self.tipo_var,
                               values=["Todos","NF-e","NFC-e","CT-e","NFSE","MDF-e"],
                               width=9, state="readonly")
        tipo_cb.pack(side="left", padx=4, pady=6)
        tipo_cb.bind("<<ComboboxSelected>>", self._trigger)

        # Status
        tk.Label(self, text="Status:", bg=COLORS["bg_card"],
                 fg=COLORS["text_secondary"], font=FONT_SMALL).pack(side="left", padx=(12,2))
        self.status_var = tk.StringVar(value="Todos")
        status_cb = ttk.Combobox(self, textvariable=self.status_var,
                                 values=["Todos","Autorizada","Cancelada","Substituída"],
                                 width=12, state="readonly")
        status_cb.pack(side="left", padx=4, pady=6)
        status_cb.bind("<<ComboboxSelected>>", self._trigger)

        # Fluxo
        tk.Label(self, text="Fluxo:", bg=COLORS["bg_card"],
                 fg=COLORS["text_secondary"], font=FONT_SMALL).pack(side="left", padx=(12,2))
        self.fluxo_var = tk.StringVar(value="Todos")
        fluxo_cb = ttk.Combobox(self, textvariable=self.fluxo_var,
                                values=["Todos","Saída Própria","Entrada de Terceiros",
                                        "Entrada (CT-e)","Saída (CT-e)",
                                        "Entrada (NFSE)","Saída (NFSE)",
                                        "CANCELADO","SUBSTITUÍDA","Terceiros"],
                                width=18, state="readonly")
        fluxo_cb.pack(side="left", padx=4, pady=6)
        fluxo_cb.bind("<<ComboboxSelected>>", self._trigger)

        # Período
        tk.Label(self, text="De:", bg=COLORS["bg_card"],
                 fg=COLORS["text_secondary"], font=FONT_SMALL).pack(side="left", padx=(12,2))
        self.data_ini_var = tk.StringVar()
        tk.Entry(self, textvariable=self.data_ini_var,
                 bg=COLORS["bg_input"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"],
                 relief="flat", font=FONT_BODY, width=10,
                 ).pack(side="left", padx=2, pady=6)

        tk.Label(self, text="Até:", bg=COLORS["bg_card"],
                 fg=COLORS["text_secondary"], font=FONT_SMALL).pack(side="left", padx=(4,2))
        self.data_fim_var = tk.StringVar()
        tk.Entry(self, textvariable=self.data_fim_var,
                 bg=COLORS["bg_input"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"],
                 relief="flat", font=FONT_BODY, width=10,
                 ).pack(side="left", padx=2, pady=6)

        # Só inconsistências
        self.inc_var = tk.BooleanVar(value=False)
        tk.Checkbutton(self, text="Só Inconsistências",
                       variable=self.inc_var, command=self._trigger,
                       bg=COLORS["bg_card"], fg=COLORS["text_secondary"],
                       selectcolor=COLORS["bg_input"], activebackground=COLORS["bg_card"],
                       font=FONT_SMALL).pack(side="left", padx=12)

        # Limpar filtros
        StyledButton(self, text="Limpar Filtros", style="secondary",
                     command=self._clear_filters).pack(side="left", padx=4)

    def _trigger(self, *args):
        if self.on_filter:
            self.on_filter()

    def _clear_filters(self):
        self.search_var.set("")
        self.tipo_var.set("Todos")
        self.status_var.set("Todos")
        self.fluxo_var.set("Todos")
        self.data_ini_var.set("")
        self.data_fim_var.set("")
        self.inc_var.set(False)
        self._trigger()

    def get_filters(self):
        return {
            "search":     self.search_var.get().lower().strip(),
            "tipo":       self.tipo_var.get(),
            "status":     self.status_var.get(),
            "fluxo":      self.fluxo_var.get(),
            "data_ini":   self.data_ini_var.get().strip(),
            "data_fim":   self.data_fim_var.get().strip(),
            "inc_only":   self.inc_var.get(),
        }


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  App Principal                                                               ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"FiscalPro Analytics  v{APP_VERSION}  —  NF-e / CT-e / NFSe (padrão nacional) / MDF-e")
        self.geometry("1600x960")
        self.minsize(1200, 700)
        self.configure(bg=COLORS["bg_main"])

        # Ícone da janela (skip se não disponível)
        try:
            self.iconbitmap("icon.ico")
        except Exception:
            pass

        # ─── Estado ──────────────────────────────────────────────────────────
        self.documentos_processados   = {}
        self.arquivos_contados        = 0
        self.cfop_totals              = defaultdict(float)
        self.nfse_item_totals         = defaultdict(lambda: {"bruto":0.,"retencoes":0.,"liquido":0.})
        self.partner_totals           = defaultdict(float)
        self.erros_detectados         = 0
        self.quebras_sequencia_alerts = []
        self._processing              = False
        self._work_queue              = queue.Queue()
        # NFSe substituição: chave substituída → chave nova
        self.substituicoes            = {}
        # Log de auditoria fiscal
        self.audit_log                = []

        GLOBAL_ITEM_DETAILS.clear()

        # ─── Config persistente ──────────────────────────────────────────────
        self._config = self._load_config()
        self.cnpj_empresa_var     = tk.StringVar(value=self._config.get("cnpj",""))
        self.nome_empresa_var     = tk.StringVar(value=self._config.get("nome_empresa",""))
        self.pasta_padrao         = self._config.get("pasta_padrao","")

        # ─── TTK Styles ──────────────────────────────────────────────────────
        self._configure_styles()

        # ─── Layout Principal ─────────────────────────────────────────────────
        self._build_layout()

    # ── Estilos TTK ──────────────────────────────────────────────────────────────
    def _configure_styles(self):
        style = ttk.Style(self)
        style.theme_use("default")

        # Treeview
        style.configure("Custom.Treeview",
            background=COLORS["bg_table"],
            foreground=COLORS["text_primary"],
            fieldbackground=COLORS["bg_table"],
            rowheight=26,
            font=FONT_SMALL,
            borderwidth=0,
        )
        style.configure("Custom.Treeview.Heading",
            background=COLORS["bg_header"],
            foreground=COLORS["text_secondary"],
            font=("Segoe UI", 9, "bold"),
            relief="flat",
            borderwidth=0,
        )
        style.map("Custom.Treeview",
            background=[("selected", COLORS["accent_dim"])],
            foreground=[("selected", COLORS["text_white"])],
        )
        style.map("Custom.Treeview.Heading",
            background=[("active", COLORS["bg_card"])],
        )

        # Scrollbar
        style.configure("Dark.Vertical.TScrollbar",
            troughcolor=COLORS["bg_main"],
            background=COLORS["bg_input"],
            bordercolor=COLORS["bg_main"],
            arrowcolor=COLORS["text_dim"],
        )
        style.configure("Dark.Horizontal.TScrollbar",
            troughcolor=COLORS["bg_main"],
            background=COLORS["bg_input"],
            bordercolor=COLORS["bg_main"],
            arrowcolor=COLORS["text_dim"],
        )

        # Progressbar
        style.configure("TProgressbar",
            troughcolor=COLORS["bg_input"],
            background=COLORS["accent"],
            bordercolor=COLORS["bg_input"],
        )

        # Combobox
        style.configure("TCombobox",
            fieldbackground=COLORS["bg_input"],
            background=COLORS["bg_input"],
            foreground=COLORS["text_primary"],
            selectbackground=COLORS["accent_dim"],
            selectforeground=COLORS["text_white"],
        )
        style.map("TCombobox",
            fieldbackground=[("readonly", COLORS["bg_input"])],
            foreground=[("readonly", COLORS["text_primary"])],
        )

        # Notebook (abas)
        style.configure("Dark.TNotebook",
            background=COLORS["bg_main"],
            borderwidth=0,
            tabmargins=0,
        )
        style.configure("Dark.TNotebook.Tab",
            background=COLORS["bg_card"],
            foreground=COLORS["text_secondary"],
            padding=[16, 8],
            font=FONT_SMALL,
            borderwidth=0,
        )
        style.map("Dark.TNotebook.Tab",
            background=[("selected", COLORS["bg_main"]),
                        ("active",   COLORS["bg_card_hover"])],
            foreground=[("selected", COLORS["text_primary"]),
                        ("active",   COLORS["text_primary"])],
        )

        # Separator
        style.configure("Dark.TSeparator",
            background=COLORS["border"],
        )

    # ── Layout ───────────────────────────────────────────────────────────────────
    def _build_layout(self):
        # Sidebar (esquerda)
        self._build_sidebar()

        # Conteúdo principal (direita)
        main = tk.Frame(self, bg=COLORS["bg_main"])
        main.pack(side="left", fill="both", expand=True)

        # Header
        self._build_header(main)

        # KPI Cards
        self._build_kpi_bar(main)

        # Notebook de abas
        self._build_notebook(main)

        # Status bar
        self.status_bar = StatusBar(main)
        self.status_bar.pack(side="bottom", fill="x")

    def _build_sidebar(self):
        sidebar = tk.Frame(self, bg=COLORS["bg_sidebar"], width=230)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        # Logo / Título
        logo_frame = tk.Frame(sidebar, bg=COLORS["bg_sidebar"], pady=20)
        logo_frame.pack(fill="x")

        tk.Label(logo_frame, text="⬡", bg=COLORS["bg_sidebar"],
                 fg=COLORS["accent"], font=("Segoe UI", 28)).pack()
        tk.Label(logo_frame, text="FiscalPro",
                 bg=COLORS["bg_sidebar"], fg=COLORS["text_white"],
                 font=("Segoe UI", 15, "bold")).pack()
        tk.Label(logo_frame, text=f"Analytics  v{APP_VERSION}",
                 bg=COLORS["bg_sidebar"], fg=COLORS["text_dim"],
                 font=FONT_SMALL).pack()

        ttk.Separator(sidebar, orient="horizontal", style="Dark.TSeparator").pack(fill="x", padx=16, pady=8)

        # CNPJ / Empresa
        fields = tk.Frame(sidebar, bg=COLORS["bg_sidebar"], padx=14)
        fields.pack(fill="x", pady=4)

        tk.Label(fields, text="EMPRESA", bg=COLORS["bg_sidebar"],
                 fg=COLORS["text_dim"], font=("Segoe UI", 8, "bold")).pack(anchor="w")
        tk.Entry(fields, textvariable=self.nome_empresa_var,
                 bg=COLORS["bg_input"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"],
                 relief="flat", font=FONT_SMALL,
                 ).pack(fill="x", pady=(2,8))

        tk.Label(fields, text="CNPJ (só números)", bg=COLORS["bg_sidebar"],
                 fg=COLORS["text_dim"], font=("Segoe UI", 8, "bold")).pack(anchor="w")
        tk.Entry(fields, textvariable=self.cnpj_empresa_var,
                 bg=COLORS["bg_input"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"],
                 relief="flat", font=FONT_SMALL,
                 ).pack(fill="x", pady=(2,10))

        ttk.Separator(sidebar, orient="horizontal", style="Dark.TSeparator").pack(fill="x", padx=16, pady=4)

        # Botões de ação
        btn_frame = tk.Frame(sidebar, bg=COLORS["bg_sidebar"], padx=14)
        btn_frame.pack(fill="x", pady=4)

        actions = [
            ("📂  Selecionar Arquivos",    "primary",   self.selecionar_e_processar_arquivos),
            ("🗂   Selecionar Pasta",        "secondary", self.selecionar_pasta),
            ("🚫  Importar Canceladas",      "danger",    self.importar_como_canceladas),
            ("🧹  Limpar Resultados",        "secondary", self.limpar_resultados),
            ("📊  Exportar XLSX",            "success",   self.exportar_xlsx),
            ("📋  Exportar CSV",             "secondary", self.exportar_csv),
            ("📄  Relatório PDF",            "secondary", self.exportar_pdf),
            ("💾  Salvar Sessão",            "secondary", self.salvar_sessao),
            ("📥  Carregar Sessão",          "secondary", self.carregar_sessao),
        ]
        for txt, sty, cmd in actions:
            StyledButton(btn_frame, text=txt, style=sty,
                         command=cmd).pack(fill="x", pady=2)

        ttk.Separator(sidebar, orient="horizontal", style="Dark.TSeparator").pack(fill="x", padx=16, pady=8)

        # Info rápida de erros/quebras
        info = tk.Frame(sidebar, bg=COLORS["bg_sidebar"], padx=14)
        info.pack(fill="x")

        tk.Label(info, text="DIAGNÓSTICO RÁPIDO", bg=COLORS["bg_sidebar"],
                 fg=COLORS["text_dim"], font=("Segoe UI", 8, "bold")).pack(anchor="w", pady=(0,6))

        self.lbl_erros_side = tk.Label(info, text="⚠  Erros: 0",
                                        bg=COLORS["bg_sidebar"],
                                        fg=COLORS["accent3"], font=FONT_SMALL)
        self.lbl_erros_side.pack(anchor="w", pady=2)

        self.lbl_quebras_side = tk.Label(info, text="🔢  Quebras: 0",
                                          bg=COLORS["bg_sidebar"],
                                          fg=COLORS["accent4"], font=FONT_SMALL)
        self.lbl_quebras_side.pack(anchor="w", pady=2)

        self.lbl_incons_side = tk.Label(info, text="🚩  Inconsistências: 0",
                                          bg=COLORS["bg_sidebar"],
                                          fg=COLORS["accent4"], font=FONT_SMALL)
        self.lbl_incons_side.pack(anchor="w", pady=2)

        StyledButton(info, text="Ver Quebras de Seq.", style="warning",
                     command=self.mostrar_quebras_popup).pack(fill="x", pady=6)

        # Rodapé
        tk.Label(sidebar, text="© 2025 FiscalPro",
                 bg=COLORS["bg_sidebar"], fg=COLORS["text_dim"],
                 font=("Segoe UI", 8)).pack(side="bottom", pady=10)

    def _build_header(self, parent):
        header = tk.Frame(parent, bg=COLORS["bg_header"], height=56)
        header.pack(fill="x")
        header.pack_propagate(False)

        tk.Label(header, text="Analisador Fiscal de Documentos Eletrônicos",
                 bg=COLORS["bg_header"], fg=COLORS["text_white"],
                 font=FONT_HEADING).pack(side="left", padx=20, pady=14)

        # Data/hora
        self.lbl_datetime = tk.Label(header, text="",
                                      bg=COLORS["bg_header"],
                                      fg=COLORS["text_dim"], font=FONT_SMALL)
        self.lbl_datetime.pack(side="right", padx=20)
        self._update_clock()

        # Arquivos processados
        self.lbl_arquivos = tk.Label(header, text="Arquivos: 0",
                                      bg=COLORS["bg_header"],
                                      fg=COLORS["text_secondary"], font=FONT_SMALL)
        self.lbl_arquivos.pack(side="right", padx=16)

    def _update_clock(self):
        now = datetime.now().strftime("%d/%m/%Y  %H:%M:%S")
        self.lbl_datetime.configure(text=now)
        self.after(1000, self._update_clock)

    def _build_kpi_bar(self, parent):
        kpi_frame = tk.Frame(parent, bg=COLORS["bg_main"], pady=10)
        kpi_frame.pack(fill="x", padx=16)

        kpi_defs = [
            ("kpi_total",    "R$ 0,00",      "Faturamento Total Autorizado",    COLORS["accent"]),
            ("kpi_nfe",      "0 / 0",        "NF-e / NFC-e  Auth / Canc",       COLORS["accent2"]),
            ("kpi_cte",      "0 / 0",        "CT-e  Autorizados / Cancelados",  COLORS["accent4"]),
            ("kpi_nfse",     "0 / 0",        "NFSe  Autorizadas / Canceladas",  COLORS["accent5"]),
            ("kpi_erros",    "0",            "Erros de Processamento",           COLORS["accent3"]),
            ("kpi_incons",   "0",            "Inconsistências Fiscais",          COLORS["accent4"]),
        ]

        self._kpi_cards = {}
        for i, (key, val, title, color) in enumerate(kpi_defs):
            card = KPICard(kpi_frame, title=title, value=val, color=color)
            card.grid(row=0, column=i, padx=6, sticky="nsew")
            kpi_frame.grid_columnconfigure(i, weight=1)
            self._kpi_cards[key] = card

    def _build_notebook(self, parent):
        self.notebook = ttk.Notebook(parent, style="Dark.TNotebook")
        self.notebook.pack(fill="both", expand=True, padx=0, pady=0)

        tabs = [
            ("📋  Documentos",        self._build_tab_documentos),
            ("📦  Itens / Produtos",   self._build_tab_itens),
            ("🏷  CFOP",              self._build_tab_cfop),
            ("🔧  NFSe Agregado",     self._build_tab_nfse),
            ("🤝  Parceiros",         self._build_tab_parceiros),
            ("📅  Por Período",       self._build_tab_periodo),
            ("🔍  Auditoria Fiscal",  self._build_tab_auditoria),
            ("🧮  Simples Nacional",  self._build_tab_simples),
        ]

        for name, builder in tabs:
            frame = tk.Frame(self.notebook, bg=COLORS["bg_main"])
            self.notebook.add(frame, text=name)
            builder(frame)

    # ── Tabs ─────────────────────────────────────────────────────────────────────

    def _build_tab_documentos(self, parent):
        # Filtros
        self.filter_bar = FilterBar(parent, on_filter=self.aplicar_filtros)
        self.filter_bar.pack(fill="x", padx=0)

        # Contador filtrado
        count_frame = tk.Frame(parent, bg=COLORS["bg_main"])
        count_frame.pack(fill="x", padx=12, pady=(4,0))
        self.lbl_count_docs = tk.Label(count_frame, text="0 documentos",
                                        bg=COLORS["bg_main"],
                                        fg=COLORS["text_dim"], font=FONT_SMALL)
        self.lbl_count_docs.pack(side="left")

        # Tabela principal
        tree_frame = tk.Frame(parent, bg=COLORS["bg_table"])
        tree_frame.pack(fill="both", expand=True, padx=0, pady=4)

        cols = ("Chave","Tipo","Nº","Data","Valor","Status","Fluxo","Parceiro","CNPJ/CPF","Inconsistência","Arquivo")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings",
                                  style="Custom.Treeview")

        col_conf = [
            ("Chave",          180, "w"),
            ("Tipo",            52, "center"),
            ("Nº",              64, "center"),
            ("Data",            82, "center"),
            ("Valor",          110, "e"),
            ("Status",          80, "center"),
            ("Fluxo",          145, "center"),
            ("Parceiro",       200, "w"),
            ("CNPJ/CPF",       125, "center"),
            ("Inconsistência", 160, "center"),
            ("Arquivo",        180, "w"),
        ]
        for col, width, anchor in col_conf:
            self.tree.heading(col, text=col,
                              command=lambda c=col: self._sort_tree(self.tree, c, False))
            self.tree.column(col, width=width, anchor=anchor, minwidth=40)

        # Tags de cor
        self.tree.tag_configure("cancelada",    foreground=COLORS["status_cancel"])
        self.tree.tag_configure("substituida",  foreground="#C77DFF")
        self.tree.tag_configure("inconsistente",foreground=COLORS["status_warn"])
        self.tree.tag_configure("ok",           foreground=COLORS["text_primary"])
        self.tree.tag_configure("alt",          background=COLORS["bg_table_alt"])

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                             command=self.tree.yview, style="Dark.Vertical.TScrollbar")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal",
                             command=self.tree.xview, style="Dark.Horizontal.TScrollbar")
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Menu de contexto
        self._context_menu = tk.Menu(self, tearoff=0, bg=COLORS["bg_card"],
                                      fg=COLORS["text_primary"])
        self._context_menu.add_command(label="📋  Copiar Chave",
                                        command=self._copy_chave)
        self._context_menu.add_command(label="🔍  Ver Itens do Documento",
                                        command=self._ver_itens_doc)
        self._context_menu.add_separator()
        self._context_menu.add_command(label="🚩  Marcar como Revisado",
                                        command=self._marcar_revisado)
        self.tree.bind("<Button-3>", self._show_context_menu)
        self.tree.bind("<Double-1>", self._ver_itens_doc)

    def _build_tab_itens(self, parent):
        # Busca itens
        search_frame = tk.Frame(parent, bg=COLORS["bg_card"])
        search_frame.pack(fill="x")
        tk.Label(search_frame, text="🔍  Filtrar produto/serviço:",
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"],
                 font=FONT_SMALL).pack(side="left", padx=10, pady=6)
        self.item_search_var = tk.StringVar()
        self.item_search_var.trace_add("write", lambda *a: self.aplicar_filtros())
        tk.Entry(search_frame, textvariable=self.item_search_var,
                 bg=COLORS["bg_input"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"],
                 relief="flat", font=FONT_SMALL, width=30).pack(side="left", padx=4)

        tree_frame = tk.Frame(parent, bg=COLORS["bg_table"])
        tree_frame.pack(fill="both", expand=True)

        cols = ("Nº Doc","Produto / Serviço","NCM","CEST","Qtd","Un","Valor Unit","Valor Total","CST/CSOSN","PIS/COF CST","vICMS","Sugestão Tributação")
        self.tree_itens = ttk.Treeview(tree_frame, columns=cols, show="headings",
                                        style="Custom.Treeview")
        col_conf = [
            ("Nº Doc",              60, "center"),
            ("Produto / Serviço",  270, "w"),
            ("NCM",                 88, "center"),
            ("CEST",                70, "center"),
            ("Qtd",                 55, "e"),
            ("Un",                  45, "center"),
            ("Valor Unit",          95, "e"),
            ("Valor Total",        105, "e"),
            ("CST/CSOSN",           72, "center"),
            ("PIS/COF CST",         72, "center"),
            ("vICMS",               80, "e"),
            ("Sugestão Tributação", 400, "w"),
        ]
        for col, width, anchor in col_conf:
            self.tree_itens.heading(col, text=col)
            self.tree_itens.column(col, width=width, anchor=anchor)

        self.tree_itens.tag_configure("alt",   background=COLORS["bg_table_alt"])
        self.tree_itens.tag_configure("sugest_warn",  foreground=COLORS["accent4"])
        self.tree_itens.tag_configure("sugest_error", foreground=COLORS["accent3"])
        self.tree_itens.tag_configure("sugest_alt_warn",  foreground=COLORS["accent4"], background=COLORS["bg_table_alt"])
        self.tree_itens.tag_configure("sugest_alt_error", foreground=COLORS["accent3"], background=COLORS["bg_table_alt"])

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                             command=self.tree_itens.yview, style="Dark.Vertical.TScrollbar")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal",
                             command=self.tree_itens.xview, style="Dark.Horizontal.TScrollbar")
        self.tree_itens.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree_itens.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

    def _build_tab_cfop(self, parent):
        tree_frame = tk.Frame(parent, bg=COLORS["bg_table"])
        tree_frame.pack(fill="both", expand=True, padx=16, pady=16)

        cols = ("CFOP","Descrição","Qtd Docs","Valor Total")
        self.tree_cfop = ttk.Treeview(tree_frame, columns=cols, show="headings",
                                       style="Custom.Treeview", height=20)
        col_conf = [
            ("CFOP",       80, "center"),
            ("Descrição",  400, "w"),
            ("Qtd Docs",   80, "center"),
            ("Valor Total",150, "e"),
        ]
        for col, width, anchor in col_conf:
            self.tree_cfop.heading(col, text=col,
                                    command=lambda c=col: self._sort_tree(self.tree_cfop, c, False))
            self.tree_cfop.column(col, width=width, anchor=anchor)

        self.tree_cfop.tag_configure("alt", background=COLORS["bg_table_alt"])

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                             command=self.tree_cfop.yview, style="Dark.Vertical.TScrollbar")
        self.tree_cfop.configure(yscrollcommand=vsb.set)
        self.tree_cfop.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

    def _build_tab_nfse(self, parent):
        tree_frame = tk.Frame(parent, bg=COLORS["bg_table"])
        tree_frame.pack(fill="both", expand=True, padx=16, pady=16)

        cols = ("Item Lista","CNAE","Cód. Trib. Mun.","Qtd","Valor Bruto","Retenções","Valor Líquido")
        self.tree_nfse = ttk.Treeview(tree_frame, columns=cols, show="headings",
                                       style="Custom.Treeview", height=20)
        col_conf = [
            ("Item Lista",     130, "center"),
            ("CNAE",           110, "center"),
            ("Cód. Trib. Mun.",150, "center"),
            ("Qtd",             60, "center"),
            ("Valor Bruto",    140, "e"),
            ("Retenções",      140, "e"),
            ("Valor Líquido",  140, "e"),
        ]
        for col, width, anchor in col_conf:
            self.tree_nfse.heading(col, text=col)
            self.tree_nfse.column(col, width=width, anchor=anchor)

        self.tree_nfse.tag_configure("alt", background=COLORS["bg_table_alt"])

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                             command=self.tree_nfse.yview, style="Dark.Vertical.TScrollbar")
        self.tree_nfse.configure(yscrollcommand=vsb.set)
        self.tree_nfse.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

    def _build_tab_parceiros(self, parent):
        tree_frame = tk.Frame(parent, bg=COLORS["bg_table"])
        tree_frame.pack(fill="both", expand=True, padx=16, pady=16)

        cols = ("Parceiro","CNPJ/CPF","Qtd Docs","Valor Total","Último Doc")
        self.tree_partner = ttk.Treeview(tree_frame, columns=cols, show="headings",
                                          style="Custom.Treeview", height=20)
        col_conf = [
            ("Parceiro",  300, "w"),
            ("CNPJ/CPF",  140, "center"),
            ("Qtd Docs",   80, "center"),
            ("Valor Total",150, "e"),
            ("Último Doc",  90, "center"),
        ]
        for col, width, anchor in col_conf:
            self.tree_partner.heading(col, text=col,
                                       command=lambda c=col: self._sort_tree(self.tree_partner, c, False))
            self.tree_partner.column(col, width=width, anchor=anchor)

        self.tree_partner.tag_configure("alt", background=COLORS["bg_table_alt"])

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                             command=self.tree_partner.yview, style="Dark.Vertical.TScrollbar")
        self.tree_partner.configure(yscrollcommand=vsb.set)
        self.tree_partner.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

    def _build_tab_periodo(self, parent):
        info = tk.Frame(parent, bg=COLORS["bg_main"], pady=20)
        info.pack(fill="both", expand=True)

        tk.Label(info, text="📅  Resumo por Mês de Competência",
                 bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                 font=FONT_HEADING).pack(anchor="w", padx=20, pady=(0,10))

        tree_frame = tk.Frame(info, bg=COLORS["bg_table"])
        tree_frame.pack(fill="both", expand=True, padx=16)

        cols = ("Período","NF-e Auth","NF-e Canc","CT-e Auth","CT-e Canc",
                "NFSe Auth","NFSe Canc","Valor Total Auth","Valor Total Canc")
        self.tree_periodo = ttk.Treeview(tree_frame, columns=cols, show="headings",
                                          style="Custom.Treeview", height=20)
        for col in cols:
            self.tree_periodo.heading(col, text=col)
            w = 140 if "Valor" in col else 80
            self.tree_periodo.column(col, width=w, anchor="center")

        self.tree_periodo.tag_configure("alt", background=COLORS["bg_table_alt"])

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                             command=self.tree_periodo.yview, style="Dark.Vertical.TScrollbar")
        self.tree_periodo.configure(yscrollcommand=vsb.set)
        self.tree_periodo.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

    # ─────────────────────────────────────────────────────────────────────────────
    # ABA SIMPLES NACIONAL — Pré-apuração com memória de cálculo
    # ─────────────────────────────────────────────────────────────────────────────
    # Tabelas LC 155/2016 (vigente 2018+)
    _SN_ANEXOS = {
        "III": {
            "desc": "Anexo III — Prestação de Serviços",
            "faixas": [
                (180000,      0.060, 0),
                (360000,      0.112, 9360),
                (720000,      0.135, 17640),
                (1800000,     0.160, 35640),
                (3600000,     0.210, 125640),
                (4800000,     0.330, 648000),
            ],
            "particip": {"IRPJ":4.00,"CSLL":3.50,"COFINS":13.64,"PIS":2.96,"CPP":43.40,"ISS":32.50},
        },
        "V": {
            "desc": "Anexo V — Serviços (com fator r)",
            "faixas": [
                (180000,      0.155, 0),
                (360000,      0.180, 4500),
                (720000,      0.195, 9900),
                (1800000,     0.205, 17100),
                (3600000,     0.230, 62100),
                (4800000,     0.305, 540000),
            ],
            "particip": {"IRPJ":25.00,"CSLL":15.00,"COFINS":14.10,"PIS":3.05,"CPP":28.85,"ISS":14.00},
        },
        "I": {
            "desc": "Anexo I — Comércio",
            "faixas": [
                (180000,      0.040, 0),
                (360000,      0.073, 5940),
                (720000,      0.095, 13860),
                (1800000,     0.107, 22500),
                (3600000,     0.143, 87300),
                (4800000,     0.190, 378000),
            ],
            "particip": {"IRPJ":5.50,"CSLL":3.50,"COFINS":12.74,"PIS":2.76,"CPP":41.50,"ICMS":34.00},
        },
        "II": {
            "desc": "Anexo II — Indústria",
            "faixas": [
                (180000,      0.045, 0),
                (360000,      0.078, 5940),
                (720000,      0.100, 13860),
                (1800000,     0.112, 22500),
                (3600000,     0.147, 85500),
                (4800000,     0.300, 720000),
            ],
            "particip": {"IRPJ":5.50,"CSLL":3.50,"COFINS":11.51,"PIS":2.49,"CPP":37.50,"IPI":7.50,"ICMS":32.00},
        },
    }

    def _build_tab_simples(self, parent):
        """Aba de pré-apuração do Simples Nacional com memória de cálculo."""
        outer = tk.Frame(parent, bg=COLORS["bg_main"])
        outer.pack(fill="both", expand=True)

        # ── Inputs ───────────────────────────────────────────────────────────
        inp = tk.Frame(outer, bg=COLORS["bg_card"])
        inp.pack(fill="x", padx=0)

        tk.Label(inp, text="🧮  Calculadora Simples Nacional — Pré-Apuração",
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                 font=FONT_HEADING).grid(row=0, column=0, columnspan=8,
                                         padx=14, pady=(10,4), sticky="w")

        # Anexo
        tk.Label(inp, text="Anexo:", bg=COLORS["bg_card"],
                 fg=COLORS["text_secondary"], font=FONT_LABEL).grid(row=1,column=0,padx=(14,4),pady=6,sticky="e")
        self.sn_anexo_var = tk.StringVar(value="III")
        anexo_cb = ttk.Combobox(inp, textvariable=self.sn_anexo_var,
                                values=list(self._SN_ANEXOS.keys()), width=6, state="readonly")
        anexo_cb.grid(row=1,column=1,padx=4,pady=6,sticky="w")
        anexo_cb.bind("<<ComboboxSelected>>", lambda e: self._calcular_simples())

        # RPA — será preenchido automaticamente mas pode editar
        tk.Label(inp, text="RPA (R$):", bg=COLORS["bg_card"],
                 fg=COLORS["text_secondary"], font=FONT_LABEL).grid(row=1,column=2,padx=(16,4),pady=6,sticky="e")
        self.sn_rpa_var = tk.StringVar()
        tk.Entry(inp, textvariable=self.sn_rpa_var, width=14,
                 bg=COLORS["bg_input"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"], relief="flat",
                 font=FONT_BODY).grid(row=1,column=3,padx=4,pady=6,sticky="w")

        # RBT12
        tk.Label(inp, text="RBT12 (R$):", bg=COLORS["bg_card"],
                 fg=COLORS["text_secondary"], font=FONT_LABEL).grid(row=1,column=4,padx=(16,4),pady=6,sticky="e")
        self.sn_rbt12_var = tk.StringVar()
        tk.Entry(inp, textvariable=self.sn_rbt12_var, width=14,
                 bg=COLORS["bg_input"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"], relief="flat",
                 font=FONT_BODY).grid(row=1,column=5,padx=4,pady=6,sticky="w")

        # Fator r (para Anexo III/V)
        tk.Label(inp, text="Folha 12m (R$):", bg=COLORS["bg_card"],
                 fg=COLORS["text_secondary"], font=FONT_LABEL).grid(row=1,column=6,padx=(16,4),pady=6,sticky="e")
        self.sn_folha_var = tk.StringVar(value="0")
        tk.Entry(inp, textvariable=self.sn_folha_var, width=14,
                 bg=COLORS["bg_input"], fg=COLORS["text_primary"],
                 insertbackground=COLORS["text_primary"], relief="flat",
                 font=FONT_BODY).grid(row=1,column=7,padx=(4,14),pady=6,sticky="w")

        # Botões
        btn_row = tk.Frame(inp, bg=COLORS["bg_card"])
        btn_row.grid(row=2, column=0, columnspan=8, padx=14, pady=(0,8), sticky="w")
        StyledButton(btn_row, text="⚡  Calcular",         style="primary",
                     command=self._calcular_simples).pack(side="left", padx=(0,6))
        StyledButton(btn_row, text="📥  Usar Dados do Período",style="secondary",
                     command=self._preencher_dados_simples).pack(side="left", padx=(0,6))
        StyledButton(btn_row, text="📊  Exportar Memória XLSX", style="success",
                     command=self._exportar_memoria_simples).pack(side="left")

        # ── Canvas scrollável para resultado ─────────────────────────────────
        canvas = tk.Canvas(outer, bg=COLORS["bg_main"], highlightthickness=0)
        vsb2   = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview,
                                style="Dark.Vertical.TScrollbar")
        canvas.configure(yscrollcommand=vsb2.set)
        vsb2.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True)

        self.sn_result_frame = tk.Frame(canvas, bg=COLORS["bg_main"])
        self._sn_canvas_win  = canvas.create_window((0,0), window=self.sn_result_frame, anchor="nw")
        self.sn_result_frame.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
            lambda e: canvas.itemconfig(self._sn_canvas_win, width=e.width))
        self._sn_canvas = canvas

    def _preencher_dados_simples(self):
        """Preenche RPA automaticamente com o faturamento calculado das notas."""
        total = 0.0
        for doc in self.documentos_processados.values():
            if doc["status"] == "autorizada" and "Entrada" not in doc.get("fluxo",""):
                total += doc.get("valor", 0.0)
        self.sn_rpa_var.set(f"{total:.2f}".replace(".",","))
        self.status_bar.set_message(
            f"RPA preenchido com faturamento das notas: R$ {self._fmt(total)}")

    def _parse_valor_br(self, s):
        """Converte '99.556,72' ou '99556.72' para float."""
        try:
            s = str(s).strip().replace(" ","")
            if "," in s and "." in s:
                s = s.replace(".","").replace(",",".")
            elif "," in s:
                s = s.replace(",",".")
            return float(s)
        except:
            return 0.0

    def _calcular_simples(self):
        """Executa o cálculo do Simples Nacional e renderiza a memória."""
        for w in self.sn_result_frame.winfo_children():
            w.destroy()

        anexo_key = self.sn_anexo_var.get()
        anexo     = self._SN_ANEXOS.get(anexo_key)
        if not anexo:
            return

        rpa   = self._parse_valor_br(self.sn_rpa_var.get())
        rbt12 = self._parse_valor_br(self.sn_rbt12_var.get())
        folha = self._parse_valor_br(self.sn_folha_var.get())

        if rpa <= 0 or rbt12 <= 0:
            tk.Label(self.sn_result_frame,
                     text="⚠  Preencha RPA e RBT12 para calcular.",
                     bg=COLORS["bg_main"], fg=COLORS["accent4"],
                     font=FONT_BODY).pack(padx=20, pady=20)
            return

        # Determinar faixa
        faixa_idx = 0
        for i, (limite, _, _) in enumerate(anexo["faixas"]):
            if rbt12 <= limite:
                faixa_idx = i
                break
            faixa_idx = i

        _, aliq_nom, parcela = anexo["faixas"][faixa_idx]
        faixa_min = anexo["faixas"][faixa_idx-1][0] if faixa_idx > 0 else 0
        faixa_max = anexo["faixas"][faixa_idx][0]

        # Alíquota efetiva
        aliq_efet = ((rbt12 * aliq_nom) - parcela) / rbt12
        sn_total  = rpa * aliq_efet

        # Fator r (Anexo III vs V)
        fator_r = (folha / rbt12) if rbt12 > 0 else 0.0
        fator_r_pct = fator_r * 100
        usa_fator_r = anexo_key in ("III","V")

        # Partilha por tributo
        particip = anexo["particip"]
        tributos = {}
        for trib, pct in particip.items():
            aliq_trib = aliq_efet * (pct / 100)
            valor_trib = rpa * aliq_trib
            tributos[trib] = {"pct_part": pct, "aliq_efet": aliq_trib, "valor": valor_trib}

        # ── Render ───────────────────────────────────────────────────────────
        P = self.sn_result_frame
        bg = COLORS["bg_main"]

        def label(parent, text, font=FONT_BODY, fg=COLORS["text_primary"],
                  bg2=None, anchor="w", padx=0, pady=0):
            tk.Label(parent, text=text, bg=bg2 or bg, fg=fg,
                     font=font, anchor=anchor).pack(anchor=anchor, padx=padx, pady=pady)

        def section(title):
            f = tk.Frame(P, bg=COLORS["bg_card"],
                         highlightbackground=COLORS["border"], highlightthickness=1)
            f.pack(fill="x", padx=16, pady=6)
            tk.Label(f, text=title, bg=COLORS["bg_card"],
                     fg=COLORS["accent"], font=FONT_LABEL,
                     padx=12, pady=6).pack(anchor="w")
            return f

        def row_table(parent, cols, is_header=False, bg2=None):
            r = tk.Frame(parent, bg=bg2 or COLORS["bg_table"])
            r.pack(fill="x")
            widths = [200,180,180,180,180]
            for ci, (col, w) in enumerate(zip(cols, widths)):
                fg2 = COLORS["text_white"] if is_header else (
                      COLORS["accent"]     if ci > 0 else COLORS["text_primary"])
                tk.Label(r, text=col, bg=bg2 or COLORS["bg_table"],
                         fg=fg2 if not is_header else COLORS["text_secondary"],
                         font=FONT_LABEL if is_header else FONT_MONO,
                         width=w//8, anchor="e" if ci > 0 else "w",
                         padx=8, pady=5).pack(side="left")
            return r

        # ── Título ────────────────────────────────────────────────────────────
        empresa = self.nome_empresa_var.get() or "Empresa"
        cnpj_e  = self.cnpj_empresa_var.get()
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

        top_f = tk.Frame(P, bg=COLORS["bg_card"])
        top_f.pack(fill="x", padx=16, pady=(12,4))
        tk.Label(top_f, text=f"SIMPLES NACIONAL — PRÉ-APURAÇÃO",
                 bg=COLORS["bg_card"], fg=COLORS["accent"],
                 font=FONT_TITLE).pack(anchor="w", padx=12, pady=(8,2))
        tk.Label(top_f, text=f"{empresa}  |  CNPJ: {cnpj_e}  |  Emitido: {now_str}",
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"],
                 font=FONT_SMALL).pack(anchor="w", padx=12, pady=(0,8))

        # ── Discriminativo de Receitas ────────────────────────────────────────
        s1 = section("📊  Discriminativo de Receitas")
        dados_receita = [
            ("Receita Bruta do Período (RPA):",         f"R$ {self._fmt(rpa)}"),
            ("RBT12 — Receita dos últimos 12 meses:",   f"R$ {self._fmt(rbt12)}"),
            ("Faixa de enquadramento:",
             f"R$ {self._fmt(faixa_min)} a R$ {self._fmt(faixa_max)}"),
            ("Alíquota nominal da faixa:",               f"{aliq_nom*100:.2f}%"),
            ("Parcela a deduzir:",                       f"R$ {self._fmt(parcela)}"),
        ]
        if usa_fator_r:
            dados_receita.append(("Folha de pagamento 12m:", f"R$ {self._fmt(folha)}"))
            dados_receita.append(("Fator r (Folha/RBT12):",  f"{fator_r_pct:.2f}%"))

        for k, v in dados_receita:
            row = tk.Frame(s1, bg=COLORS["bg_card"])
            row.pack(fill="x", padx=0)
            tk.Label(row, text=k, bg=COLORS["bg_card"], fg=COLORS["text_secondary"],
                     font=FONT_BODY, width=45, anchor="w", padx=14).pack(side="left")
            tk.Label(row, text=v, bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                     font=FONT_BODY, anchor="w").pack(side="left")

        # ── Memória de Cálculo ────────────────────────────────────────────────
        s2 = section("🔢  Memória de Cálculo — Alíquota Efetiva")
        calc_rows = [
            ("RBT12 × Alíq. Nominal:", f"R$ {self._fmt(rbt12)} × {aliq_nom*100:.2f}% = R$ {self._fmt(rbt12*aliq_nom)}"),
            ("− Parcela a Deduzir:",   f"R$ {self._fmt(rbt12*aliq_nom)} − R$ {self._fmt(parcela)} = R$ {self._fmt(rbt12*aliq_nom - parcela)}"),
            ("÷ RBT12:",               f"R$ {self._fmt(rbt12*aliq_nom - parcela)} ÷ R$ {self._fmt(rbt12)}"),
            ("= Alíquota Efetiva:",    f"{aliq_efet*100:.10f}%  →  {aliq_efet*100:.4f}%"),
        ]
        for k, v in calc_rows:
            row = tk.Frame(s2, bg=COLORS["bg_card"])
            row.pack(fill="x")
            lbl_v_color = COLORS["accent"] if "Alíquota Efetiva" in k else COLORS["text_primary"]
            lbl_font    = FONT_LABEL if "Alíquota Efetiva" in k else FONT_MONO
            tk.Label(row, text=k, bg=COLORS["bg_card"], fg=COLORS["text_secondary"],
                     font=FONT_BODY, width=28, anchor="w", padx=14).pack(side="left")
            tk.Label(row, text=v, bg=COLORS["bg_card"], fg=lbl_v_color,
                     font=lbl_font, anchor="w").pack(side="left")

        # ── Repartição por Tributo ────────────────────────────────────────────
        s3 = section(f"📋  Repartição — {anexo['desc']}")

        # Header
        cols_h = ["Tributo","% Repartição","Alíq. Efetiva","Base de Cálculo","Valor (R$)"]
        h_row = tk.Frame(s3, bg=COLORS["bg_header"])
        h_row.pack(fill="x")
        col_widths = [80,130,130,160,130]
        for ci, (col, w) in enumerate(zip(cols_h, col_widths)):
            tk.Label(h_row, text=col, bg=COLORS["bg_header"],
                     fg=COLORS["text_secondary"], font=FONT_LABEL,
                     width=w//8, anchor="e" if ci > 0 else "w",
                     padx=8, pady=6).pack(side="left")

        total_check = 0.0
        for idx, (trib, dados) in enumerate(tributos.items()):
            row_bg = COLORS["bg_table_alt"] if idx % 2 == 0 else COLORS["bg_table"]
            r = tk.Frame(s3, bg=row_bg)
            r.pack(fill="x")
            total_check += dados["valor"]
            vals = [
                trib,
                f"{dados['pct_part']:.2f}%",
                f"{dados['aliq_efet']*100:.6f}%",
                f"R$ {self._fmt(rpa)}",
                f"R$ {self._fmt(dados['valor'])}",
            ]
            trib_colors = {
                "IRPJ":  "#A78BFA","CSLL":  "#C77DFF",
                "COFINS":"#60A5FA","PIS":   "#93C5FD",
                "CPP":   "#4ADE80","ISS":   "#00D4AA",
                "ICMS":  "#FB923C","IPI":   "#F87171",
            }
            for ci, (val, w) in enumerate(zip(vals, col_widths)):
                fg2 = trib_colors.get(trib, COLORS["text_primary"]) if ci == 0 else (
                      COLORS["accent"] if ci == 4 else COLORS["text_primary"])
                tk.Label(r, text=val, bg=row_bg, fg=fg2,
                         font=FONT_LABEL if ci in (0,4) else FONT_MONO,
                         width=w//8, anchor="e" if ci > 0 else "w",
                         padx=8, pady=5).pack(side="left")

        # Linha de total
        r_total = tk.Frame(s3, bg=COLORS["accent_dim"])
        r_total.pack(fill="x")
        tk.Label(r_total, text="SIMPLES NACIONAL A RECOLHER",
                 bg=COLORS["accent_dim"], fg=COLORS["text_white"],
                 font=FONT_LABEL, padx=14, pady=8).pack(side="left")
        tk.Label(r_total, text=f"R$ {self._fmt(sn_total)}",
                 bg=COLORS["accent_dim"], fg=COLORS["accent2"],
                 font=("Segoe UI",14,"bold"), padx=14, pady=8).pack(side="right")

        # ── Alerta fator r ────────────────────────────────────────────────────
        if usa_fator_r and rbt12 > 0:
            alerta_frame = tk.Frame(P, bg=COLORS["bg_card"],
                                    highlightbackground=COLORS["accent4"], highlightthickness=1)
            alerta_frame.pack(fill="x", padx=16, pady=4)
            if fator_r >= 0.28:
                msg = f"✅  Fator r = {fator_r_pct:.2f}% ≥ 28% → Tributação pelo Anexo III"
                fg3 = COLORS["accent2"]
            else:
                msg = f"⚠️  Fator r = {fator_r_pct:.2f}% < 28% → Tributação pelo Anexo V (mais oneroso)"
                fg3 = COLORS["accent4"]
            tk.Label(alerta_frame, text=msg, bg=COLORS["bg_card"], fg=fg3,
                     font=FONT_BODY, padx=14, pady=8).pack(anchor="w")

        # ── Projeção próximo período ──────────────────────────────────────────
        s4 = section("📅  Projeção Alíquota Período Seguinte")
        # Substitui o mês mais antigo pelo RPA atual (janela deslizante de 12 meses)
        mes_mais_antigo = rbt12 / 12  # estimativa do mês saindo da janela
        rbt12_next = rbt12 - mes_mais_antigo + rpa
        aliq_efet_next = None
        for i, (limite, an, pd) in enumerate(anexo["faixas"]):
            if rbt12_next <= limite:
                aliq_efet_next = ((rbt12_next * an) - pd) / rbt12_next
                break
        if aliq_efet_next is None:
            aliq_efet_next = ((rbt12_next * anexo["faixas"][-1][1]) - anexo["faixas"][-1][2]) / rbt12_next

        proj_rows = [
            ("RBT12 projetado (aprox.):",          f"R$ {self._fmt(rbt12_next)}"),
            ("Alíquota efetiva projetada:",         f"{aliq_efet_next*100:.4f}%"),
            ("Simples projetado (base = RPA atual):", f"R$ {self._fmt(rpa * aliq_efet_next)}"),
        ]
        for k, v in proj_rows:
            row = tk.Frame(s4, bg=COLORS["bg_card"])
            row.pack(fill="x")
            tk.Label(row, text=k, bg=COLORS["bg_card"], fg=COLORS["text_secondary"],
                     font=FONT_BODY, width=45, anchor="w", padx=14).pack(side="left")
            tk.Label(row, text=v, bg=COLORS["bg_card"], fg=COLORS["accent"],
                     font=FONT_BODY, anchor="w").pack(side="left")

        # Guarda resultado para exportação
        self._sn_last_result = {
            "empresa": empresa, "cnpj": cnpj_e, "now": now_str,
            "anexo_key": anexo_key, "anexo_desc": anexo["desc"],
            "rpa": rpa, "rbt12": rbt12, "folha": folha,
            "faixa_min": faixa_min, "faixa_max": faixa_max,
            "aliq_nom": aliq_nom, "parcela": parcela,
            "aliq_efet": aliq_efet, "sn_total": sn_total,
            "tributos": tributos, "fator_r": fator_r,
            "rbt12_next": rbt12_next, "aliq_efet_next": aliq_efet_next,
        }

    def _exportar_memoria_simples(self):
        """Exporta a memória de cálculo do Simples Nacional para XLSX."""
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("XLSX", "openpyxl não instalado.")
            return
        res = getattr(self, "_sn_last_result", None)
        if not res:
            messagebox.showwarning("Simples Nacional",
                                   "Calcule primeiro antes de exportar.")
            return
        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")],
            title="Salvar Memória de Cálculo SN",
        )
        if not caminho:
            return
        try:
            wb   = openpyxl.Workbook()
            ws   = wb.active
            ws.title = "Simples Nacional"
            ws.sheet_view.showGridLines = False

            C_BG    = PatternFill("solid", fgColor="0D1117")
            C_CARD  = PatternFill("solid", fgColor="1A1D2E")
            C_ACC   = PatternFill("solid", fgColor="4F7EFF")
            C_ALT   = PatternFill("solid", fgColor="161927")
            C_TOTAL = PatternFill("solid", fgColor="2A4499")

            F_TITLE  = Font("Segoe UI", size=14, bold=True, color="E8EAF6")
            F_HEAD   = Font("Segoe UI", size=9,  bold=True, color="FFFFFF")
            F_KEY    = Font("Segoe UI", size=9,  bold=False,color="8B90C4")
            F_VAL    = Font("Segoe UI", size=9,  bold=False,color="E8EAF6")
            F_ACCENT = Font("Segoe UI", size=10, bold=True, color="4F7EFF")
            F_TOTAL  = Font("Segoe UI", size=12, bold=True, color="00D4AA")

            ws.column_dimensions["A"].width = 45
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 22
            ws.column_dimensions["D"].width = 22
            ws.column_dimensions["E"].width = 22

            row = 1
            def wcell(r,c,v,font=None,fill=None,align=None):
                cell = ws.cell(r,c,v)
                if font:  cell.font  = font
                if fill:  cell.fill  = fill
                if align: cell.alignment = align
                return cell

            # Título
            ws.merge_cells(f"A{row}:E{row}")
            wcell(row,1,f"SIMPLES NACIONAL — PRÉ-APURAÇÃO",F_TITLE,C_BG)
            ws.row_dimensions[row].height = 30; row+=1

            ws.merge_cells(f"A{row}:E{row}")
            wcell(row,1,f"{res['empresa']}  |  CNPJ: {res['cnpj']}  |  {res['now']}",
                  F_KEY,C_BG); row+=1

            ws.merge_cells(f"A{row}:E{row}")
            wcell(row,1,res['anexo_desc'],F_ACCENT,C_CARD); row+=2

            # Discriminativo
            ws.merge_cells(f"A{row}:E{row}")
            wcell(row,1,"DISCRIMINATIVO DE RECEITAS",F_HEAD,C_ACC)
            ws.row_dimensions[row].height = 20; row+=1

            disc = [
                ("RPA — Receita Bruta do Período",  f"R$ {self._fmt(res['rpa'])}"),
                ("RBT12 — Últimos 12 meses",         f"R$ {self._fmt(res['rbt12'])}"),
                ("Faixa de enquadramento",
                 f"R$ {self._fmt(res['faixa_min'])} a R$ {self._fmt(res['faixa_max'])}"),
                ("Alíquota nominal",                 f"{res['aliq_nom']*100:.2f}%"),
                ("Parcela a deduzir",                f"R$ {self._fmt(res['parcela'])}"),
            ]
            if res["folha"] > 0:
                disc.append(("Folha de pagamento 12m", f"R$ {self._fmt(res['folha'])}"))
                disc.append(("Fator r",               f"{res['fator_r']*100:.2f}%"))

            for k, v in disc:
                bg_row = C_CARD if row % 2 == 0 else C_ALT
                wcell(row,1,k,F_KEY,bg_row)
                wcell(row,2,v,F_VAL,bg_row,Alignment(horizontal="right"))
                ws.row_dimensions[row].height = 18; row+=1

            row+=1

            # Memória
            ws.merge_cells(f"A{row}:E{row}")
            wcell(row,1,"MEMÓRIA DE CÁLCULO — ALÍQUOTA EFETIVA",F_HEAD,C_ACC)
            ws.row_dimensions[row].height = 20; row+=1

            rbt12 = res["rbt12"]; aliq_nom = res["aliq_nom"]; parcela = res["parcela"]
            mem = [
                (f"RBT12 × Alíquota nominal",
                 f"R$ {self._fmt(rbt12)} × {aliq_nom*100:.2f}% = R$ {self._fmt(rbt12*aliq_nom)}"),
                ("− Parcela a deduzir",
                 f"R$ {self._fmt(rbt12*aliq_nom)} − R$ {self._fmt(parcela)} = R$ {self._fmt(rbt12*aliq_nom-parcela)}"),
                ("÷ RBT12",
                 f"÷ R$ {self._fmt(rbt12)}"),
                ("= Alíquota Efetiva",
                 f"{res['aliq_efet']*100:.10f}%"),
            ]
            for k, v in mem:
                bg_row = C_CARD if row % 2 == 0 else C_ALT
                is_total = "Efetiva" in k
                wcell(row,1,k, F_ACCENT if is_total else F_KEY, bg_row)
                wcell(row,2,v, F_ACCENT if is_total else F_VAL, bg_row,
                      Alignment(horizontal="right"))
                ws.row_dimensions[row].height = 18; row+=1

            row+=1

            # Repartição
            ws.merge_cells(f"A{row}:E{row}")
            wcell(row,1,"REPARTIÇÃO POR TRIBUTO",F_HEAD,C_ACC)
            ws.row_dimensions[row].height = 20; row+=1

            for col, hdr in enumerate(["Tributo","% Repartição","Alíq. Efetiva","Base de Cálculo","Valor (R$)"],1):
                wcell(row,col,hdr,F_HEAD,C_ACC,Alignment(horizontal="center"))
            ws.row_dimensions[row].height = 20; row+=1

            for idx,(trib,dados) in enumerate(res["tributos"].items()):
                bg_row = C_ALT if idx%2==0 else C_CARD
                wcell(row,1,trib,                              F_VAL,bg_row)
                wcell(row,2,f"{dados['pct_part']:.2f}%",      F_VAL,bg_row,Alignment(horizontal="right"))
                wcell(row,3,f"{dados['aliq_efet']*100:.6f}%", F_VAL,bg_row,Alignment(horizontal="right"))
                wcell(row,4,res["rpa"],                        F_VAL,bg_row,Alignment(horizontal="right"))
                ws.cell(row,4).number_format = "#,##0.00"
                wcell(row,5,dados["valor"],                    F_ACCENT,bg_row,Alignment(horizontal="right"))
                ws.cell(row,5).number_format = "#,##0.00"
                ws.row_dimensions[row].height = 18; row+=1

            # Total
            wcell(row,1,"SIMPLES NACIONAL A RECOLHER",F_TOTAL,C_TOTAL)
            ws.merge_cells(f"A{row}:D{row}")
            wcell(row,5,res["sn_total"],F_TOTAL,C_TOTAL,Alignment(horizontal="right"))
            ws.cell(row,5).number_format = "#,##0.00"
            ws.row_dimensions[row].height = 26; row+=2

            # Próximo período
            ws.merge_cells(f"A{row}:E{row}")
            wcell(row,1,"PROJEÇÃO PERÍODO SEGUINTE",F_HEAD,C_ACC)
            ws.row_dimensions[row].height = 20; row+=1
            proj = [
                ("RBT12 projetado (janela 12m)",       f"R$ {self._fmt(res['rbt12_next'])}"),
                ("Alíquota efetiva projetada",        f"{res['aliq_efet_next']*100:.4f}%"),
                ("Simples projetado (base = RPA atual)",
                 f"R$ {self._fmt(res['rpa'] * res['aliq_efet_next'])}"),
            ]
            for k, v in proj:
                bg_row = C_CARD if row%2==0 else C_ALT
                wcell(row,1,k,F_KEY,bg_row)
                wcell(row,2,v,F_VAL,bg_row,Alignment(horizontal="right"))
                ws.row_dimensions[row].height=18; row+=1

            wb.save(caminho)
            messagebox.showinfo("Simples Nacional",
                                f"Memória de cálculo exportada:\n{caminho}")
        except Exception as e:
            logging.exception(f"Erro exportar SN: {e}")
            messagebox.showerror("Simples Nacional", f"Erro: {e}")

    def _build_tab_auditoria(self, parent):
        """Aba de auditoria fiscal — substituições, inconsistências, quebras, conciliação."""
        # ── Painel superior: resumo de auditoria ─────────────────────────────
        top = tk.Frame(parent, bg=COLORS["bg_card"])
        top.pack(fill="x")

        tk.Label(top, text="🔍  Log de Auditoria Fiscal e Conciliação Mensal",
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                 font=FONT_HEADING).pack(anchor="w", padx=14, pady=(10,4))

        btn_row = tk.Frame(top, bg=COLORS["bg_card"])
        btn_row.pack(fill="x", padx=14, pady=(0,8))
        StyledButton(btn_row, text="🔄  Atualizar Auditoria",  style="primary",
                     command=self._atualizar_auditoria).pack(side="left", padx=(0,6))
        StyledButton(btn_row, text="📋  Exportar Log CSV",     style="secondary",
                     command=self._exportar_audit_csv).pack(side="left", padx=(0,6))
        StyledButton(btn_row, text="🔢  Ver Quebras de Seq.",  style="warning",
                     command=self.mostrar_quebras_popup).pack(side="left", padx=(0,6))

        # ── Notebook interno de sub-abas ──────────────────────────────────────
        self.audit_notebook = ttk.Notebook(parent, style="Dark.TNotebook")
        self.audit_notebook.pack(fill="both", expand=True)

        # Sub-aba 1: Log geral
        f1 = tk.Frame(self.audit_notebook, bg=COLORS["bg_main"])
        self.audit_notebook.add(f1, text="📋  Log de Eventos")
        self._build_audit_log_tree(f1)

        # Sub-aba 2: Conciliação mensal (faturamento vs. documentos)
        f2 = tk.Frame(self.audit_notebook, bg=COLORS["bg_main"])
        self.audit_notebook.add(f2, text="📊  Conciliação Mensal")
        self._build_audit_conciliacao(f2)

        # Sub-aba 3: Checklist Simples Nacional
        f3 = tk.Frame(self.audit_notebook, bg=COLORS["bg_main"])
        self.audit_notebook.add(f3, text="✅  Checklist Simples Nacional")
        self._build_audit_checklist(f3)

        # Sub-aba 4: Substituições
        f4 = tk.Frame(self.audit_notebook, bg=COLORS["bg_main"])
        self.audit_notebook.add(f4, text="🔁  Substituições NFSe")
        self._build_audit_substituicoes(f4)

    def _build_audit_log_tree(self, parent):
        tree_frame = tk.Frame(parent, bg=COLORS["bg_table"])
        tree_frame.pack(fill="both", expand=True, padx=0, pady=4)

        cols = ("Data","Tipo Evento","Chave","Ref. Chave","Descrição","Arquivo")
        self.tree_audit_log = ttk.Treeview(tree_frame, columns=cols, show="headings",
                                            style="Custom.Treeview")
        col_w = [90,140,80,80,380,180]
        for col, w in zip(cols, col_w):
            self.tree_audit_log.heading(col, text=col)
            self.tree_audit_log.column(col, width=w, anchor="w")
        self.tree_audit_log.tag_configure("subst",  foreground="#C77DFF")
        self.tree_audit_log.tag_configure("error",  foreground=COLORS["accent3"])
        self.tree_audit_log.tag_configure("info",   foreground=COLORS["accent2"])
        self.tree_audit_log.tag_configure("warn",   foreground=COLORS["accent4"])
        self.tree_audit_log.tag_configure("alt",    background=COLORS["bg_table_alt"])

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                             command=self.tree_audit_log.yview, style="Dark.Vertical.TScrollbar")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal",
                             command=self.tree_audit_log.xview, style="Dark.Horizontal.TScrollbar")
        self.tree_audit_log.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree_audit_log.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

    def _build_audit_conciliacao(self, parent):
        info = tk.Frame(parent, bg=COLORS["bg_main"])
        info.pack(fill="x", padx=14, pady=8)
        tk.Label(info,
                 text="Conciliação: totais por mês separados por Saída (faturamento) vs. Entrada (compras/serviços tomados).",
                 bg=COLORS["bg_main"], fg=COLORS["text_secondary"],
                 font=FONT_SMALL).pack(anchor="w")

        tree_frame = tk.Frame(parent, bg=COLORS["bg_table"])
        tree_frame.pack(fill="both", expand=True, padx=0)

        cols = ("Mês","Docs Saída","Valor Saída","Docs Entrada","Valor Entrada",
                "Docs Serviços","Valor Serviços","Total Geral","Substituídas","Canceladas")
        self.tree_conciliacao = ttk.Treeview(tree_frame, columns=cols, show="headings",
                                              style="Custom.Treeview")
        col_w = [80,90,130,90,130,90,130,140,90,90]
        for col, w in zip(cols, col_w):
            self.tree_conciliacao.heading(col, text=col)
            self.tree_conciliacao.column(col, width=w, anchor="center")
        self.tree_conciliacao.tag_configure("alt", background=COLORS["bg_table_alt"])
        self.tree_conciliacao.tag_configure("total",
            background=COLORS["accent_dim"], foreground=COLORS["text_white"])

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                             command=self.tree_conciliacao.yview, style="Dark.Vertical.TScrollbar")
        self.tree_conciliacao.configure(yscrollcommand=vsb.set)
        self.tree_conciliacao.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

    def _build_audit_checklist(self, parent):
        """Checklist automático de obrigações Simples Nacional."""
        canvas = tk.Canvas(parent, bg=COLORS["bg_main"], highlightthickness=0)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview,
                             style="Dark.Vertical.TScrollbar")
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True)

        self.checklist_frame = tk.Frame(canvas, bg=COLORS["bg_main"])
        canvas.create_window((0,0), window=self.checklist_frame, anchor="nw")
        self.checklist_frame.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        # Conteúdo preenchido por _atualizar_auditoria
        self._checklist_canvas = canvas

    def _build_audit_substituicoes(self, parent):
        tree_frame = tk.Frame(parent, bg=COLORS["bg_table"])
        tree_frame.pack(fill="both", expand=True)

        cols = ("Nota Substituída","Nota Substituta","Motivo","Data","Valor","Arquivo")
        self.tree_substituicoes = ttk.Treeview(tree_frame, columns=cols, show="headings",
                                                style="Custom.Treeview")
        col_w = [200,200,260,90,110,180]
        for col, w in zip(cols, col_w):
            self.tree_substituicoes.heading(col, text=col)
            self.tree_substituicoes.column(col, width=w, anchor="w")
        self.tree_substituicoes.tag_configure("subst", foreground="#C77DFF")
        self.tree_substituicoes.tag_configure("alt",   background=COLORS["bg_table_alt"])

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                             command=self.tree_substituicoes.yview, style="Dark.Vertical.TScrollbar")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal",
                             command=self.tree_substituicoes.xview, style="Dark.Horizontal.TScrollbar")
        self.tree_substituicoes.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree_substituicoes.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

    def _atualizar_auditoria(self):
        """Preenche todas as sub-abas de auditoria com dados atuais."""
        if not hasattr(self, "tree_audit_log"):
            return

        # ── Log de Auditoria ─────────────────────────────────────────────────
        for i in self.tree_audit_log.get_children(): self.tree_audit_log.delete(i)

        # Gera eventos a partir dos documentos
        eventos = list(self.audit_log)  # eventos já registrados (substituições, etc.)

        # Adiciona inconsistências como eventos
        for chave, doc in self.documentos_processados.items():
            inc = doc.get("inconsistencia_fiscal","OK")
            if inc not in ("OK","","CANCELADA") and not inc.startswith("SUBSTITUÍDA"):
                eventos.append({
                    "tipo": "INCONSISTÊNCIA FISCAL",
                    "chave": chave,
                    "chave_ref": "",
                    "descricao": inc,
                    "arquivo": doc.get("arquivo",""),
                    "data": doc.get("data_emissao",""),
                })
            if doc["status"] == "cancelada":
                eventos.append({
                    "tipo": "CANCELAMENTO",
                    "chave": chave,
                    "chave_ref": "",
                    "descricao": f"{doc.get('tipo','')} nº{doc.get('nNF','')} — CANCELADA",
                    "arquivo": doc.get("arquivo",""),
                    "data": doc.get("data_emissao",""),
                })

        # Quebras de sequência
        for alerta in self.quebras_sequencia_alerts:
            eventos.append({
                "tipo": "QUEBRA DE SEQUÊNCIA",
                "chave": "", "chave_ref": "",
                "descricao": alerta,
                "arquivo": "", "data": "",
            })

        tag_map = {
            "NFSe SUBSTITUIÇÃO":  "subst",
            "INCONSISTÊNCIA FISCAL": "warn",
            "CANCELAMENTO":       "info",
            "QUEBRA DE SEQUÊNCIA":"error",
        }
        for idx, ev in enumerate(sorted(eventos, key=lambda x: x.get("data",""), reverse=True)):
            tag = tag_map.get(ev["tipo"], "alt")
            if idx % 2 == 0 and tag == "alt":
                tag = "alt"
            self.tree_audit_log.insert("", "end", tags=(tag,), values=(
                ev.get("data",""),
                ev.get("tipo",""),
                str(ev.get("chave",""))[-20:],
                str(ev.get("chave_ref",""))[-20:],
                ev.get("descricao",""),
                ev.get("arquivo",""),
            ))

        # ── Conciliação Mensal ────────────────────────────────────────────────
        for i in self.tree_conciliacao.get_children(): self.tree_conciliacao.delete(i)

        meses = defaultdict(lambda: {
            "saida_qtd":0,"saida_val":0.,
            "entrada_qtd":0,"entrada_val":0.,
            "serv_qtd":0,"serv_val":0.,
            "subst":0,"cancel":0,
        })
        for doc in self.documentos_processados.values():
            data = doc.get("data_emissao","")
            mes  = data[5:7] + "/" + data[:4] if len(data) >= 7 else "Sem data"
            st   = doc["status"]
            fluxo = doc.get("fluxo","")
            val   = doc.get("valor",0.)
            tipo  = doc.get("tipo","")

            if st == "cancelada":
                meses[mes]["cancel"] += 1
            elif st == "substituida":
                meses[mes]["subst"] += 1
            elif st == "autorizada":
                if tipo == "NFSE":
                    meses[mes]["serv_qtd"] += 1
                    meses[mes]["serv_val"] += val
                elif "Saída" in fluxo:
                    meses[mes]["saida_qtd"] += 1
                    meses[mes]["saida_val"] += val
                elif "Entrada" in fluxo:
                    meses[mes]["entrada_qtd"] += 1
                    meses[mes]["entrada_val"] += val
                else:
                    meses[mes]["serv_qtd"] += 1
                    meses[mes]["serv_val"] += val

        total_s = total_e = total_sv = 0.
        alt = False
        for mes in sorted(meses.keys()):
            m = meses[mes]
            total_geral = m["saida_val"] + m["entrada_val"] + m["serv_val"]
            total_s  += m["saida_val"]
            total_e  += m["entrada_val"]
            total_sv += m["serv_val"]
            tags = ("alt",) if alt else ()
            alt  = not alt
            self.tree_conciliacao.insert("", "end", tags=tags, values=(
                mes,
                m["saida_qtd"],   self._fmt(m["saida_val"]),
                m["entrada_qtd"], self._fmt(m["entrada_val"]),
                m["serv_qtd"],    self._fmt(m["serv_val"]),
                self._fmt(total_geral),
                m["subst"], m["cancel"],
            ))
        # Linha de total
        self.tree_conciliacao.insert("", "end", tags=("total",), values=(
            "TOTAL", "—", self._fmt(total_s),
            "—", self._fmt(total_e),
            "—", self._fmt(total_sv),
            self._fmt(total_s + total_e + total_sv),
            sum(m["subst"] for m in meses.values()),
            sum(m["cancel"] for m in meses.values()),
        ))

        # ── Checklist Simples Nacional ────────────────────────────────────────
        for widget in self.checklist_frame.winfo_children():
            widget.destroy()

        tk.Label(self.checklist_frame,
                 text="Checklist Automático — Simples Nacional",
                 bg=COLORS["bg_main"], fg=COLORS["text_primary"],
                 font=FONT_HEADING).pack(anchor="w", padx=16, pady=(12,8))

        n_inc = sum(1 for d in self.documentos_processados.values()
                    if d.get("inconsistencia_fiscal","OK") not in ("OK","CANCELADA",""))
        n_subst = len(self.substituicoes)
        n_cancel = sum(1 for d in self.documentos_processados.values() if d["status"]=="cancelada")
        n_quebras = len(self.quebras_sequencia_alerts)
        total_docs = len(self.documentos_processados)

        checks = [
            (n_inc == 0,       f"Sem inconsistências fiscais ({n_inc} encontradas)",
             "Verificar CST/CSOSN, CFOP, retenções indevidas"),
            (n_quebras == 0,   f"Sem quebras de numeração ({n_quebras} alertas)",
             "Verificar sequência numérica das NF-e e NFSe"),
            (n_subst == 0,     f"Sem NFSe substituídas ({n_subst} substituições)",
             "Verificar se substituições foram necessárias e justificadas"),
            (True,             f"Total de cancelamentos: {n_cancel}",
             "Conferir se cancelamentos foram lançados corretamente no PGDAS"),
            (total_docs > 0,   f"Total de documentos processados: {total_docs}",
             "Confrontar com relatório do emissor / prefeitura"),
            (True,             "Confrontar base de cálculo do PGDAS com faturamento apurado acima",
             "Soma de NFSe autorizadas não substituídas + NF-e saídas"),
            (True,             "Verificar retenções na fonte recebidas",
             "ISS retido só é válido em casos específicos para optantes do SN"),
            (True,             "Conferir competência dos documentos emitidos",
             "Data de emissão deve coincidir com o mês de apuração do PGDAS"),
        ]

        for ok, label, dica in checks:
            row = tk.Frame(self.checklist_frame, bg=COLORS["bg_card"],
                           highlightbackground=COLORS["border"], highlightthickness=1)
            row.pack(fill="x", padx=16, pady=3)
            icon_c = COLORS["accent2"] if ok else COLORS["accent4"]
            icon   = "✅" if ok else "⚠️"
            tk.Label(row, text=icon, bg=COLORS["bg_card"],
                     fg=icon_c, font=("Segoe UI",11)).pack(side="left", padx=10, pady=6)
            txt_frame = tk.Frame(row, bg=COLORS["bg_card"])
            txt_frame.pack(side="left", fill="x", expand=True, pady=4)
            tk.Label(txt_frame, text=label,
                     bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                     font=FONT_BODY, anchor="w").pack(anchor="w", padx=4)
            tk.Label(txt_frame, text=f"💡 {dica}",
                     bg=COLORS["bg_card"], fg=COLORS["text_dim"],
                     font=FONT_SMALL, anchor="w").pack(anchor="w", padx=4)

        self._checklist_canvas.configure(scrollregion=self._checklist_canvas.bbox("all"))

        # ── Substituições ─────────────────────────────────────────────────────
        for i in self.tree_substituicoes.get_children(): self.tree_substituicoes.delete(i)

        alt = False
        for ev in self.audit_log:
            if ev.get("tipo") != "NFSe SUBSTITUIÇÃO":
                continue
            ch_orig  = ev.get("chave_ref","")
            ch_nova  = ev.get("chave","")
            doc_nova = self.documentos_processados.get(ch_nova,{})
            tags = ("subst",) if not alt else ("alt",)
            alt  = not alt
            self.tree_substituicoes.insert("", "end", tags=tags, values=(
                ch_orig,
                ch_nova,
                ev.get("descricao","").split("Motivo:")[-1].strip() if "Motivo:" in ev.get("descricao","") else "—",
                ev.get("data",""),
                self._fmt(doc_nova.get("valor",0.)),
                ev.get("arquivo",""),
            ))

    def _exportar_audit_csv(self):
        if not self.audit_log and not self.documentos_processados:
            messagebox.showwarning("Auditoria", "Nenhum dado de auditoria para exportar.")
            return
        caminho = filedialog.asksaveasfilename(
            defaultextension=".csv", filetypes=[("CSV","*.csv")],
            title="Salvar Log de Auditoria CSV"
        )
        if not caminho: return
        try:
            with open(caminho,"w",newline="",encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(["Data","Tipo","Chave","Ref. Chave","Descrição","Arquivo"])
                for ev in self.audit_log:
                    w.writerow([ev.get("data",""), ev.get("tipo",""),
                                ev.get("chave",""), ev.get("chave_ref",""),
                                ev.get("descricao",""), ev.get("arquivo","")])
                # Inconsistências
                for chave, doc in self.documentos_processados.items():
                    inc = doc.get("inconsistencia_fiscal","OK")
                    if inc not in ("OK","","CANCELADA"):
                        w.writerow([doc.get("data_emissao",""), "INCONSISTÊNCIA",
                                    chave,"", inc, doc.get("arquivo","")])
            messagebox.showinfo("Auditoria", f"Log exportado:\n{caminho}")
        except Exception as e:
            messagebox.showerror("Auditoria", f"Erro: {e}")

    # ── Helpers ──────────────────────────────────────────────────────────────────

    def _find_text_any(self, root, candidates):
        for p in candidates:
            node = root.find(p)
            if node is not None and node.text and node.text.strip():
                return node.text.strip()
        return None

    def _find_node_any(self, root, candidates):
        for p in candidates:
            node = root.find(p)
            if node is not None:
                return node
        return None

    def _fmt(self, valor):
        return f"{valor:,.2f}".replace(",","X").replace(".",",").replace("X",".")

    def _sort_tree(self, tree, col, reverse):
        data = [(tree.set(k, col), k) for k in tree.get_children("")]
        try:
            data.sort(key=lambda x: float(x[0].replace(".","").replace(",",".")), reverse=reverse)
        except ValueError:
            data.sort(key=lambda x: x[0], reverse=reverse)
        for idx, (_, k) in enumerate(data):
            tree.move(k, "", idx)
        tree.heading(col, command=lambda: self._sort_tree(tree, col, not reverse))

    def _show_context_menu(self, event):
        row = self.tree.identify_row(event.y)
        if row:
            self.tree.selection_set(row)
            self._context_menu.tk_popup(event.x_root, event.y_root)

    def _copy_chave(self):
        sel = self.tree.selection()
        if sel:
            chave = self.tree.item(sel[0])["values"][0]
            self.clipboard_clear()
            self.clipboard_append(str(chave))
            self.status_bar.set_message(f"Chave copiada: {chave}")

    def _ver_itens_doc(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        chave = str(self.tree.item(sel[0])["values"][0])
        itens = GLOBAL_ITEM_DETAILS.get(chave, [])
        if not itens:
            messagebox.showinfo("Itens", f"Sem itens detalhados para: {chave}")
            return
        self.notebook.select(1)  # aba Itens

    def _marcar_revisado(self):
        sel = self.tree.selection()
        if sel:
            self.tree.item(sel[0], tags=("revisado",))
            self.tree.tag_configure("revisado", foreground=COLORS["accent2"])

    # ── Sugestão de Tributação por NCM / CEST ────────────────────────────────────
    # Produtos sujeitos à Substituição Tributária (CEST obrigatório - Conv. ICMS 92/15)
    # Formato: cest_prefixo → (descrição segmento, CSOSN_SN, CST_RN, CFOP_saida_correta)
    _CEST_ST_MAP = {
        "01": ("Autopeças",                        "201", "10", "5401"),
        "02": ("Bebidas alcoólicas exceto cerveja", "201", "10", "5401"),
        "03": ("Cervejas, chopes, refrigerantes",   "201", "10", "5401"),
        "04": ("Cigarros e outros produtos tabaco", "500", "60", "5405"),
        "05": ("Cimentos",                          "201", "10", "5401"),
        "06": ("Combustíveis e lubrificantes",      "500", "60", "5405"),
        "07": ("Energia elétrica",                  "500", "60", "5405"),
        "08": ("Ferramentas",                       "201", "10", "5401"),
        "09": ("Lâmpadas, reatores e starter",      "201", "10", "5401"),
        "10": ("Materiais de construção e congên.", "201", "10", "5401"),
        "11": ("Materiais de limpeza",              "201", "10", "5401"),
        "12": ("Materiais elétricos",               "201", "10", "5401"),
        "13": ("Medicamentos de uso humano",        "500", "60", "5405"),
        "14": ("Papéis, plásticos, produtos embal.","201", "10", "5401"),
        "15": ("Pneumáticos",                       "201", "10", "5401"),
        "16": ("Produtos alimentícios",             "201", "10", "5401"),
        "17": ("Produtos de papelaria",             "201", "10", "5401"),
        "18": ("Produtos de perfumaria",            "201", "10", "5401"),
        "19": ("Produtos eletrônicos, eletrdom.",   "201", "10", "5401"),
        "20": ("Rações para animais domésticos",    "201", "10", "5401"),
        "21": ("Sorvetes e preparados p/ sorvetes", "201", "10", "5401"),
        "22": ("Tintas e vernizes",                 "201", "10", "5401"),
        "23": ("Veículos automotores",              "201", "10", "5401"),
        "24": ("Veículos de duas e três rodas",     "201", "10", "5401"),
        "25": ("Vending machines",                  "201", "10", "5401"),
        "26": ("Bicicletas",                        "201", "10", "5401"),
        "27": ("Cosméticos, perfumaria, higiene",   "201", "10", "5401"),
        "28": ("Bebidas quentes",                   "201", "10", "5401"),
    }

    # Produtos com PIS/COFINS Monofásico (Lei 10.485/02 e 10.147/00)
    # NCM prefixo → (descrição, CST_monofasico)
    _NCM_PISCOFINS_MONO = {
        # Farmacêuticos (Lei 10.147/00 art 1º)
        "3001": ("Farmacêuticos - glândulas/extratos",   "02"),
        "3002": ("Farmacêuticos - sangue/soros",         "02"),
        "3003": ("Medicamentos - misturas p/ terapia",   "02"),
        "3004": ("Medicamentos - doses",                 "02"),
        "3005": ("Algodão/gazes/ataduras médicas",       "02"),
        "3006": ("Preparações farmacêuticas",            "02"),
        # Perfumaria/Higiene (Lei 10.147/00 art 1º)
        "3303": ("Perfumes e águas de colônia",          "02"),
        "3304": ("Produtos de beleza/maquiagem",         "02"),
        "3305": ("Preparações p/ cabelos",               "02"),
        "3306": ("Preparações p/ higiene bucal",         "02"),
        "3307": ("Preparações p/ barba/desodorantes",    "02"),
        "3401": ("Sabões e detergentes",                 "02"),
        "3402": ("Agentes orgânicos de superfície",      "02"),
        # Autopeças (Lei 10.485/02 art 3º)
        "4011": ("Pneus novos de borracha",              "02"),
        "4013": ("Câmaras-de-ar de borracha",            "02"),
        "7009": ("Espelhos de vidro",                    "02"),
        "8421": ("Centrifugadores/filtros",              "02"),
        "8483": ("Veios/rolamentos",                     "02"),
        "8507": ("Acumuladores elétricos (baterias)",    "02"),
        "8512": ("Equip. elétricos iluminação veículos", "02"),
        "8527": ("Rádios para veículos",                 "02"),
        "8708": ("Partes e acessórios p/ veículos",      "02"),
        "9401": ("Assentos (bancos para veículos)",      "02"),
        # Bebidas (Lei 10.833/03 art 58-A e seg)
        "2201": ("Águas/gelo",                           "02"),
        "2202": ("Refrigerantes/bebidas não alcoólicas", "02"),
        "2203": ("Cervejas de malte",                    "02"),
        "2204": ("Vinhos de uvas frescas",               "02"),
        "2205": ("Vermutes",                             "02"),
        "2206": ("Outras bebidas fermentadas",           "02"),
        "2207": ("Álcool etílico",                       "02"),
        "2208": ("Aguardentes/uísques/gins",             "02"),
        # Combustíveis (Lei 9.718/98 art 4º - monofásico)
        "2710": ("Óleos combustíveis/lubrificantes",     "02"),
        "2711": ("GLP/GN",                               "02"),
        # Veículos (Lei 10.485/02 art 1º)
        "8701": ("Tratores",                             "02"),
        "8702": ("Veículos p/ transporte 10+ pessoas",   "02"),
        "8703": ("Automóveis de passeio",                "02"),
        "8704": ("Veículos p/ transporte de mercadorias","02"),
        "8705": ("Veículos p/ usos especiais",           "02"),
        "8711": ("Motocicletas/ciclomotores",            "02"),
        "8712": ("Bicicletas",                           "03"),  # alíq. diferenciada
    }

    # CFOP correto por situação tributária ICMS
    _CFOP_CORRECAO = {
        # (tpNF, csosn_ou_cst) → cfop_correto
        # Vendas com ST cobrada pelo remetente (Simples Nacional)
        ("1", "201"): ("5401", "Venda de prod. fabricado com ST"),
        ("1", "202"): ("5401", "Venda de prod. fabricado com ST"),
        ("1", "203"): ("5402", "Venda de prod. industrializado ST"),
        # Vendas com ST já retida (comprou com ST, revende)
        ("1", "500"): ("5405", "Venda de merc. com ST retida anteriormente"),
        # Vendas normais SN sem ST
        ("1", "101"): ("5102", "Venda de merc. adquirida ou recebida de terceiros"),
        ("1", "102"): ("5102", "Venda de merc. adquirida ou recebida de terceiros"),
        ("1", "400"): ("5102", "Venda de merc. adquirida ou recebida de terceiros"),
        # Regime Normal
        ("1", "00"): ("5102", "Venda normal tributada"),
        ("1", "10"): ("5401", "Venda com ST"),
        ("1", "20"): ("5102", "Venda com redução de BC"),
        ("1", "40"): ("5102", "Venda isenta"),
        ("1", "41"): ("5102", "Venda não tributada"),
        ("1", "60"): ("5405", "Venda com ST retida"),
        ("1", "70"): ("5401", "Venda com redução BC + ST"),
    }

    def _sugerir_tributacao_ncm(self, ncm: str, cest: str, is_simples: bool,
                                 cst_icms: str, cfop_atual: str,
                                 cst_pis: str, cst_cof: str) -> str:
        """Motor de sugestão fiscal: ICMS/ST por CEST, PIS/COFINS monofásico, CFOP correto."""
        alertas = []
        ncm_clean  = (ncm  or "").replace(".", "").strip()
        cest_clean = (cest or "").replace(".", "").strip()
        cst_clean  = str(cst_icms).strip()
        cfop_clean = str(cfop_atual).strip()

        # ── 1. Verificar Substituição Tributária via CEST ─────────────────────
        tem_st = False
        cfop_sugerido = None
        if cest_clean and len(cest_clean) >= 2:
            seg = cest_clean[:2]
            if seg in self._CEST_ST_MAP:
                desc_seg, csosn_sn, cst_rn, cfop_st = self._CEST_ST_MAP[seg]
                tem_st = True

                if is_simples:
                    # Verifica CSOSN correto para ST
                    csosn_esperado_list = ("201", "202", "203", "500")
                    if cst_clean not in csosn_esperado_list:
                        alertas.append(
                            f"⚠ ST ({desc_seg}): CSOSN deveria ser 500 (ST já retida) "
                            f"ou 201/202 (ST a recolher) — atual: CSOSN {cst_clean}"
                        )
                    # Verifica CFOP para ST já retida (comprou com ST)
                    if cst_clean == "500" and cfop_clean not in ("5405","6404","7"):
                        cfop_sugerido = "5405"
                        alertas.append(
                            f"⚠ CFOP incorreto: produto com CSOSN 500 (ST retida) "
                            f"deve usar CFOP 5405 — atual: {cfop_clean}"
                        )
                    elif cst_clean in ("201","202") and cfop_clean not in ("5401","5402","6401","6402"):
                        cfop_sugerido = "5401"
                        alertas.append(
                            f"⚠ CFOP incorreto: produto com CSOSN {cst_clean} (ST a recolher) "
                            f"deve usar CFOP 5401 — atual: {cfop_clean}"
                        )
                    # Caso típico: tem CEST mas está usando CFOP de venda normal
                    if cfop_clean in ("5102","5101","6102","6101") and cst_clean not in ("101","102","400"):
                        cfop_sugerido = cfop_st
                        alertas.append(
                            f"⚠ Produto sujeito à ST ({desc_seg}, CEST {cest_clean[:2]}): "
                            f"CFOP {cfop_clean} pode estar incorreto → sugerido {cfop_st}"
                        )
                else:
                    if cst_clean not in ("10", "60", "70"):
                        alertas.append(
                            f"⚠ ST ({desc_seg}): CST deveria ser 10 (ST a recolher) "
                            f"ou 60 (ST já retida) — atual: CST {cst_clean}"
                        )
                    if cst_clean == "60" and cfop_clean not in ("5405","6404"):
                        cfop_sugerido = "5405"
                        alertas.append(
                            f"⚠ CFOP: CST 60 (ST retida) → usar CFOP 5405 — atual: {cfop_clean}"
                        )

        # ── 2. Verificar PIS/COFINS Monofásico pelo NCM ───────────────────────
        sugest_pisc = None
        for length in (4, 6, 8):
            key = ncm_clean[:length]
            if key in self._NCM_PISCOFINS_MONO:
                desc_mono, cst_mono = self._NCM_PISCOFINS_MONO[key]
                sugest_pisc = (desc_mono, cst_mono)
                break

        if sugest_pisc:
            desc_mono, cst_mono = sugest_pisc
            # CST 02 = monofásico (alíquota zero para revendedores)
            # CST 03 = alíquota zero específica
            cst_pis_clean = str(cst_pis).strip()
            cst_cof_clean = str(cst_cof).strip()
            if cst_pis_clean not in ("02", "03", "04", "06", "07", "49", "50", "99"):
                alertas.append(
                    f"💡 PIS/COFINS monofásico ({desc_mono}): "
                    f"considerar CST {cst_mono} (monofásico) para redução — "
                    f"atual PIS:{cst_pis_clean}/COF:{cst_cof_clean}"
                )

        # ── 3. Verificar NCM x CEST (obrigatoriedade de CEST) ─────────────────
        if not cest_clean and tem_st:
            alertas.append(f"⚠ CEST ausente — obrigatório para produtos sujeitos a ST (Conv. ICMS 92/15)")

        # ── 4. Verificar CST regime (SN deve usar CSOSN, não CST de 2 dígitos) ──
        if is_simples and cst_clean.isdigit() and len(cst_clean) == 2:
            alertas.append(
                f"⚠ CRT=1 (Simples Nacional): usar CSOSN (3 dígitos) "
                f"em vez de CST {cst_clean} (regime normal)"
            )

        return " | ".join(alertas) if alertas else ""

    # ── CFOP Descrições (principais) ─────────────────────────────────────────────
    CFOP_DESC = {
        "1102": "Compra p/ comercialização - Estadual",
        "1202": "Devolução de venda - Estadual",
        "1411": "Compra prod. rural - Estadual",
        "2102": "Compra p/ comercialização - Interestadual",
        "2202": "Devolução de venda - Interestadual",
        "3102": "Compra p/ comercialização - Importação",
        "5101": "Venda prod. industrializado - Estadual",
        "5102": "Venda mercadoria adquirida - Estadual",
        "5103": "Venda prod. industrializado uso/consumo - Estadual",
        "5112": "Venda merc. adq. recebida terceiros - Estadual",
        "5201": "Devolução compra prod. industrial - Estadual",
        "5202": "Devolução compra merc. adquirida - Estadual",
        "5301": "Venda prod. agropecuário - Estadual",
        "5302": "Venda merc. adq. prod. rural - Estadual",
        "5401": "Venda prod. fabricado substituição tributária - Estadual",
        "5402": "Venda prod. industrializado ST - Estadual",
        "5403": "Venda merc. adq. ST - Estadual",
        "5405": "Venda merc. ST, quando ST retida - Estadual",
        "5501": "Remessa p/ industrialização - Estadual",
        "5556": "Venda de bem imobilizado - Estadual",
        "5601": "Transferência de produção - Estadual",
        "5605": "Transferência de mercadoria - Estadual",
        "5651": "Venda de bem do ativo imobilizado - Estadual",
        "5667": "Venda de material de uso/consumo - Estadual",
        "5901": "Remessa p/ industrialização p/ conta e ordem",
        "5910": "Remessa em bonificação, doação ou brinde",
        "5920": "Remessa p/ conserto ou reparo",
        "5922": "Lançamento efetuado a título de simples faturamento",
        "5923": "Remessa p/ demonstração",
        "5924": "Remessa p/ venda fora do estabelecimento",
        "5925": "Remessa p/ exposição ou feira",
        "5929": "Lançamento efetuado em decorrência de desembaraço",
        "5949": "Outras saídas - Estadual",
        "6101": "Venda prod. industrializado - Interestadual",
        "6102": "Venda mercadoria adquirida - Interestadual",
        "6108": "Venda merc. adq. por filial - Interestadual",
        "6110": "Venda merc. adq. terceiros destino zona franca",
        "6201": "Devolução compra prod. industrial - Interestadual",
        "6202": "Devolução compra merc. adquirida - Interestadual",
        "6301": "Venda prod. agropecuário - Interestadual",
        "6401": "Venda prod. fabricado ST - Interestadual",
        "6404": "Venda merc. sujeita ST - Interestadual",
        "6410": "Venda merc. adq. em transferência com ST - Interestadual",
        "6503": "Devolução de merc. adq. ou recebida p/ utilização",
        "6551": "Venda de bem do ativo imobilizado - Interestadual",
        "6556": "Remessa p/ industrialização - Interestadual",
        "6900": "Outras saídas - Interestadual",
        "6910": "Remessa em bonificação, doação ou brinde",
        "6949": "Outras saídas - Interestadual",
        "7101": "Venda prod. industrializado - Exportação",
        "7102": "Venda mercadoria - Exportação",
        "7501": "Exportação de merc. recebidas p/ fins de exportação",
    }

    # ── Seleção de arquivos ───────────────────────────────────────────────────────

    def selecionar_e_processar_arquivos(self):
        initial = self.pasta_padrao if self.pasta_padrao else os.path.expanduser("~")
        caminhos = filedialog.askopenfilenames(
            title="Selecione XMLs, ZIPs ou RARs",
            initialdir=initial,
            filetypes=[
                ("Documentos Fiscais", "*.xml *.zip *.rar"),
                ("XML", "*.xml"), ("ZIP", "*.zip"), ("RAR", "*.rar"), ("Todos", "*.*"),
            ]
        )
        if caminhos:
            self._iniciar_processamento(list(caminhos))

    def importar_como_canceladas(self):
        """Importa XMLs/ZIPs forçando status cancelada — para notas cujo evento
        de cancelamento não está disponível mas o usuário sabe que foram canceladas."""
        initial = self.pasta_padrao if self.pasta_padrao else os.path.expanduser("~")
        caminhos = filedialog.askopenfilenames(
            title="Importar como Canceladas — Selecionar Arquivos",
            initialdir=initial,
            filetypes=[
                ("Documentos Fiscais", "*.xml *.zip *.rar"),
                ("XML", "*.xml"), ("ZIP", "*.zip"), ("Todos", "*.*"),
            ],
        )
        if not caminhos:
            return
        resp = messagebox.askyesno(
            "Confirmar",
            f"{len(caminhos)} arquivo(s) selecionado(s).\n\n"
            "Todos os documentos encontrados serão marcados como CANCELADOS "
            "independentemente do status no XML.\n\n"
            "Deseja continuar?",
        )
        if resp:
            self._iniciar_processamento(list(caminhos), forcar_cancel_todos=True)

    def selecionar_pasta(self):
        initial = self.pasta_padrao if self.pasta_padrao else os.path.expanduser("~")
        pasta = filedialog.askdirectory(title="Selecionar Pasta", initialdir=initial)
        if pasta:
            self.pasta_padrao = pasta
            caminhos = [
                os.path.join(pasta, f)
                for f in os.listdir(pasta)
                if os.path.splitext(f)[1].lower() in (".xml",".zip",".rar")
            ]
            # Detecta se o nome da pasta indica cancelamento
            nome_pasta = os.path.basename(pasta).upper()
            forcar_cancel_pasta = any(p in nome_pasta for p in
                                      ("CANCEL", "CANCELAD", "CANC_", "INUTILIZ"))
            if caminhos:
                self._iniciar_processamento(caminhos,
                                            forcar_cancel_todos=forcar_cancel_pasta)
            else:
                messagebox.showinfo("Pasta", "Nenhum arquivo XML/ZIP/RAR encontrado na pasta.")

    def _iniciar_processamento(self, caminhos, forcar_cancel_todos=False):
        if self._processing:
            messagebox.showwarning("Processamento", "Aguarde o processamento atual terminar.")
            return
        self.limpar_resultados(clear_files=False)
        self._processing = True
        self.status_bar.set_message(f"Processando {len(caminhos)} arquivo(s)...")
        self.status_bar.set_progress(0)

        t = threading.Thread(target=self._processar_thread,
                             args=(caminhos, forcar_cancel_todos), daemon=True)
        t.start()
        self.after(100, self._check_queue)

    def _processar_thread(self, caminhos, forcar_cancel_todos=False):
        total = len(caminhos)
        _CANCEL_KEYWORDS = ("CANCEL", "CANCELAD", "CANC_", "_CANC", "INUTILIZ")
        for idx, caminho in enumerate(caminhos):
            try:
                ext = os.path.splitext(caminho)[1].lower()
                nome_base = os.path.basename(caminho).upper()
                forcar = forcar_cancel_todos or any(p in nome_base for p in _CANCEL_KEYWORDS)
                if ext == ".xml":
                    with open(caminho, "rb") as f:
                        self._work_queue.put(("xml", f.read(),
                                              os.path.basename(caminho), forcar))
                elif ext == ".zip":
                    with zipfile.ZipFile(caminho, "r") as zf:
                        for nome in zf.namelist():
                            if nome.lower().endswith(".xml"):
                                forcar_int = forcar or any(
                                    p in nome.upper() for p in _CANCEL_KEYWORDS)
                                self._work_queue.put(("xml", zf.read(nome),
                                                      f"{os.path.basename(caminho)}/{nome}",
                                                      forcar_int))
                elif ext == ".rar" and RARFILE_AVAILABLE:
                    import rarfile as rf_mod
                    with rf_mod.RarFile(caminho, "r") as rf:
                        for nome in rf.namelist():
                            if nome.lower().endswith(".xml"):
                                forcar_int = forcar or any(
                                    p in nome.upper() for p in _CANCEL_KEYWORDS)
                                self._work_queue.put(("xml", rf.read(nome),
                                                      f"{os.path.basename(caminho)}/{nome}",
                                                      forcar_int))
                elif ext == ".rar" and not RARFILE_AVAILABLE:
                    self._work_queue.put(("erro", None,
                                          f"RAR não suportado (instale rarfile): {caminho}",
                                          False))
            except Exception as e:
                self._work_queue.put(("erro", None, f"Erro container {caminho}: {e}", False))
                logging.exception(f"Erro ao processar {caminho}: {e}")

            progress = int((idx + 1) / total * 100)
            self._work_queue.put(("progress", progress, ""))

        self._work_queue.put(("done", None, ""))

    def _check_queue(self):
        try:
            while True:
                item = self._work_queue.get_nowait()
                cmd  = item[0]
                data = item[1]
                name = item[2]
                forcar_cancel = item[3] if len(item) > 3 else False
                if cmd == "xml":
                    self.arquivos_contados += 1
                    self.processar_conteudo_xml(data, name, forcar_cancel=forcar_cancel)
                elif cmd == "erro":
                    self._registrar_erro(name)
                elif cmd == "progress":
                    self.status_bar.set_progress(data)
                elif cmd == "done":
                    self._processing = False
                    self.status_bar.set_progress(100)
                    # ── Resolve substituições pendentes ──────────────────────
                    # Caso: nota substituta foi processada ANTES da original
                    for key, val in list(self.substituicoes.items()):
                        if key.startswith("_pendente_"):
                            ch_orig = key[len("_pendente_"):]
                            chave_nova, nNF_nova, motivo = val
                            if ch_orig in self.documentos_processados:
                                d = self.documentos_processados[ch_orig]
                                d["status"] = "substituida"
                                d["inconsistencia_fiscal"] = f"SUBSTITUÍDA por nNFSe={nNF_nova}"
                                d["fluxo"] = "SUBSTITUÍDA"
                                GLOBAL_ITEM_DETAILS.pop(ch_orig, None)
                                # Remove pendente, registra par real
                                del self.substituicoes[key]
                                self.substituicoes[ch_orig] = chave_nova
                    # ─────────────────────────────────────────────────────────
                    self.atualizar_interface()
                    self._salvar_config()
                    n = len(self.documentos_processados)
                    self.status_bar.set_message(
                        f"Concluído: {self.arquivos_contados} arquivo(s), {n} documento(s) processado(s)."
                    )
                    self.status_bar.set_progress(0)
                    break   # sai do while, não da função — garante que fila seja drenada
        except queue.Empty:
            pass

        if self._processing:
            self.after(100, self._check_queue)

    # ── Parser ───────────────────────────────────────────────────────────────────

    def processar_conteudo_xml(self, conteudo_xml, nome_arquivo, forcar_cancel=False):
        try:
            cnpj_empresa = (self.cnpj_empresa_var.get()
                            .replace(".",  "").replace("/","").replace("-","").strip())

            try:
                root = ET.fromstring(conteudo_xml)
            except ET.ParseError as e:
                logging.warning(f"XML inválido {nome_arquivo}: {e}")
                self._registrar_erro(f"XML inválido {nome_arquivo}: {e}")
                return

            chave = None
            valor = 0.0
            tipo  = "N/A"
            status = "autorizada"
            nNF   = "N/A"
            data_emissao = ""
            cfop_breakdown = []
            nfse_item_breakdown = None
            cnpj_parceiro  = "N/A"
            nome_parceiro  = "N/A"
            cnpj_emitente  = "N/A"
            inconsistencia_fiscal = "OK"
            tpNF  = None
            fluxo = "N/A"
            itens_lista = []
            is_cancellation_event = False

            # ── NF-e ─────────────────────────────────────────────────────────
            infNFe = root.find(".//{http://www.portalfiscal.inf.br/nfe}infNFe")
            infCte = root.find(".//{http://www.portalfiscal.inf.br/cte}infCte")
            infMDFe= root.find(".//{http://www.portalfiscal.inf.br/mdfe}infMDFe")

            NNS = "http://www.portalfiscal.inf.br/nfe"
            CTS = "http://www.portalfiscal.inf.br/cte"
            MDS = "http://www.portalfiscal.inf.br/mdfe"

            def nfe(tag): return f"{{{NNS}}}{tag}"
            def cte(tag): return f"{{{CTS}}}{tag}"
            def mdf(tag): return f"{{{MDS}}}{tag}"

            if infNFe is not None:
                # ── Detecta modelo: 55=NF-e, 65=NFC-e (Nota Fiscal de Consumidor Eletrônica) ──
                mod_t = root.find(f".//{nfe('ide')}/{nfe('mod')}")
                mod_val = mod_t.text.strip() if mod_t is not None and mod_t.text else "55"
                tipo = "NFC-e" if mod_val == "65" else "NF-e"

                chave = infNFe.get("Id", "").replace("NFe", "") or None

                nNF_t = root.find(f".//{nfe('ide')}/{nfe('nNF')}")
                nNF   = nNF_t.text if nNF_t is not None else "N/A"

                dEmi_t = root.find(f".//{nfe('ide')}/{nfe('dhEmi')}")
                if dEmi_t is None: dEmi_t = root.find(f".//{nfe('ide')}/{nfe('dEmi')}")
                data_emissao = (dEmi_t.text or "")[:10] if dEmi_t is not None else ""

                emit_t = root.find(f".//{nfe('emit')}")
                if emit_t is not None:
                    c = emit_t.find(nfe("CNPJ"))
                    cnpj_emitente = c.text if c is not None else "N/A"
                    # Para NFC-e: parceiro = emitente (venda ao consumidor, dest pode ser ausente)
                    xNome_emit = emit_t.find(nfe("xNome"))
                    nome_emit_str = xNome_emit.text if xNome_emit is not None else ""
                else:
                    nome_emit_str = ""

                tpNF_t = root.find(f".//{nfe('ide')}/{nfe('tpNF')}")
                tpNF   = tpNF_t.text if tpNF_t is not None else None

                finNFe_t = root.find(f".//{nfe('ide')}/{nfe('finNFe')}")
                fin_nfe = finNFe_t.text if finNFe_t is not None else "1"

                # Regime tributário do emitente (para validação CST vs CSOSN)
                # indRatISSQN ou CRT: 1=SN, 2=SN-excesso, 3=regime normal
                crt_t = root.find(f".//{nfe('emit')}/{nfe('CRT')}")
                crt_val = crt_t.text if crt_t is not None else "3"
                is_simples = crt_val in ("1", "2")

                vNF_t = root.find(f".//{nfe('ICMSTot')}/{nfe('vNF')}")
                try:
                    valor = float(vNF_t.text) if vNF_t is not None and vNF_t.text else 0.0
                except:
                    valor = 0.0

                dest_t = root.find(f".//{nfe('dest')}")
                if dest_t is not None:
                    cp = dest_t.find(nfe("CNPJ"))
                    if cp is None: cp = dest_t.find(nfe("CPF"))
                    cnpj_parceiro = cp.text if cp is not None else "N/A"
                    np_t = dest_t.find(nfe("xNome"))
                    nome_parceiro = np_t.text if np_t is not None else "N/A"
                elif mod_val == "65":
                    # NFC-e ao consumidor final — dest é opcional
                    cnpj_parceiro = "CONSUMIDOR"
                    nome_parceiro = "Consumidor Final"

                nfe_flags = []
                for det_t in root.findall(f".//{nfe('det')}"):
                    # Busca prod e imposto com namespace explícito E como fallback wildcard
                    prod_t = det_t.find(nfe("prod"))
                    if prod_t is None: prod_t = det_t.find("{*}prod")
                    if prod_t is None: prod_t = det_t.find("prod")
                    imp_t  = det_t.find(nfe("imposto"))
                    if imp_t is None: imp_t = det_t.find("{*}imposto")
                    if imp_t is None: imp_t = det_t.find("imposto")
                    if prod_t is None:
                        continue

                    def _get(node, tag):
                        """Busca tag com namespace NF-e, wildcard e sem namespace.
                        USA 'is not None' para evitar o bug de bool(element)==False
                        em elementos folha (sem filhos) do ElementTree."""
                        for t in (nfe(tag), f"{{*}}{tag}", tag):
                            el = node.find(t)
                            if el is not None and el.text and el.text.strip():
                                return el.text.strip()
                        return None

                    xProd    = _get(prod_t, "xProd")    or "N/A"
                    ncm      = _get(prod_t, "NCM")       or "N/A"
                    cfop_val = _get(prod_t, "CFOP")      or "N/A"
                    cest     = _get(prod_t, "CEST")      or ""
                    cean     = _get(prod_t, "cEAN")      or ""
                    cProd    = _get(prod_t, "cProd")     or ""
                    uCom     = _get(prod_t, "uCom")      or ""
                    try:
                        qCom   = float(_get(prod_t, "qCom")   or "0")
                    except:
                        qCom   = 0.0
                    try:
                        vUnCom = float(_get(prod_t, "vUnCom") or "0")
                    except:
                        vUnCom = 0.0
                    try:
                        vProd  = float(_get(prod_t, "vProd")  or "0")
                    except:
                        vProd  = 0.0

                    cst_icms = "N/A"
                    v_icms   = 0.0
                    cst_pis  = "N/A"
                    cst_cof  = "N/A"
                    v_pis    = 0.0
                    v_cof    = 0.0
                    if imp_t is not None:
                        icms_g = imp_t.find(nfe("ICMS"))
                        if icms_g is None: icms_g = imp_t.find("{*}ICMS")
                        if icms_g is None: icms_g = imp_t.find("ICMS")
                        if icms_g is not None:
                            for child in icms_g:
                                # CST (regime normal) ou CSOSN (Simples Nacional)
                                for tag_cst in (nfe("CST"), "{*}CST", "CST",
                                                nfe("CSOSN"), "{*}CSOSN", "CSOSN"):
                                    el = child.find(tag_cst)
                                    if el is not None and el.text:
                                        cst_icms = el.text.strip(); break
                                for tag_vi in (nfe("vICMS"), "{*}vICMS", "vICMS"):
                                    el = child.find(tag_vi)
                                    if el is not None and el.text:
                                        try: v_icms = float(el.text)
                                        except: pass
                                        break
                                break
                        # PIS
                        pis_g = imp_t.find(nfe("PIS"))
                        if pis_g is None: pis_g = imp_t.find("{*}PIS")
                        if pis_g is None: pis_g = imp_t.find("PIS")
                        if pis_g is not None:
                            for child in pis_g:
                                el_cst = child.find(nfe("CST"))
                                if el_cst is None: el_cst = child.find("{*}CST")
                                if el_cst is None: el_cst = child.find("CST")
                                if el_cst is not None and el_cst.text:
                                    cst_pis = el_cst.text.strip()
                                el_v = child.find(nfe("vPIS"))
                                if el_v is None: el_v = child.find("{*}vPIS")
                                if el_v is None: el_v = child.find("vPIS")
                                if el_v is not None and el_v.text:
                                    try: v_pis = float(el_v.text)
                                    except: pass
                                break
                        # COFINS
                        cof_g = imp_t.find(nfe("COFINS"))
                        if cof_g is None: cof_g = imp_t.find("{*}COFINS")
                        if cof_g is None: cof_g = imp_t.find("COFINS")
                        if cof_g is not None:
                            for child in cof_g:
                                el_cst = child.find(nfe("CST"))
                                if el_cst is None: el_cst = child.find("{*}CST")
                                if el_cst is None: el_cst = child.find("CST")
                                if el_cst is not None and el_cst.text:
                                    cst_cof = el_cst.text.strip()
                                el_v = child.find(nfe("vCOFINS"))
                                if el_v is None: el_v = child.find("{*}vCOFINS")
                                if el_v is None: el_v = child.find("vCOFINS")
                                if el_v is not None and el_v.text:
                                    try: v_cof = float(el_v.text)
                                    except: pass
                                break

                    # Sugestão de tributação baseada no NCM+CEST (CRT do emitente)
                    sugestao_trib = self._sugerir_tributacao_ncm(ncm, cest, is_simples, cst_icms, cfop_val, cst_pis, cst_cof)

                    itens_lista.append({
                        "nNF": nNF, "nItem": det_t.get("nItem"),
                        "xProd": xProd, "NCM": ncm, "CFOP": cfop_val,
                        "CEST": cest, "cProd": cProd,
                        "qCom": qCom, "uCom": uCom, "vUnCom": vUnCom,
                        "vProd": vProd, "CST_ICMS": cst_icms, "vICMS": v_icms,
                        "CST_PIS": cst_pis, "vPIS": v_pis,
                        "CST_COF": cst_cof, "vCOFINS": v_cof,
                        "sugestao_trib": sugestao_trib,
                    })

                    # ── Revisões fiscais (respeitam CRT do emitente) ──────────
                    if tpNF == "0" and cfop_val[0:1] in ["5","6","7"]:
                        if "CFOP SAÍDA em NF ENTRADA" not in nfe_flags:
                            nfe_flags.append("CFOP SAÍDA em NF ENTRADA")
                    if tpNF == "1" and cfop_val[0:1] in ["1","2","3"]:
                        if "CFOP ENTRADA em NF SAÍDA" not in nfe_flags:
                            nfe_flags.append("CFOP ENTRADA em NF SAÍDA")
                    # CST de 2 dígitos só é inconsistência se o emitente for regime NORMAL (CRT=3)
                    # Optantes do SN (CRT=1/2) usam CSOSN — CST vem de outros tributos (IPI,PIS,COFINS)
                    if (not is_simples and
                            cst_icms and cst_icms.isdigit() and
                            len(cst_icms) == 2 and cst_icms not in ("","N/A")):
                        if "CST regime normal em optante SN - usar CSOSN" not in nfe_flags:
                            nfe_flags.append("CST regime normal em optante SN - usar CSOSN")
                    if cst_icms == "500" and tpNF == "1":
                        if "CSOSN 500 em NF Saída - verificar" not in nfe_flags:
                            nfe_flags.append("CSOSN 500 em NF Saída - verificar")
                    if v_icms > 0 and cst_icms in ("101","102","103","300","400","500"):
                        if "ICMS destacado em CSOSN sem destaque" not in nfe_flags:
                            nfe_flags.append("ICMS destacado em CSOSN sem destaque")
                    if fin_nfe == "2":
                        if "NF COMPLEMENTAR — conferir soma com NF original" not in nfe_flags:
                            nfe_flags.append("NF COMPLEMENTAR — conferir soma com NF original")
                    if fin_nfe == "3":
                        if "NF AJUSTE — verificar lançamento" not in nfe_flags:
                            nfe_flags.append("NF AJUSTE — verificar lançamento")

                    cfop_breakdown.append({"cfop": cfop_val, "vProd": vProd})

                inconsistencia_fiscal = " | ".join(nfe_flags) if nfe_flags else "OK"

            # ── CT-e ─────────────────────────────────────────────────────────
            elif infCte is not None:
                tipo  = "CT-e"
                chave = infCte.get("Id", "").replace("CTe", "") or None

                nNF_t = root.find(f".//{cte('ide')}/{cte('nCT')}")
                nNF   = nNF_t.text if nNF_t is not None else "N/A"

                dEmi_t = root.find(f".//{cte('ide')}/{cte('dhEmi')}")
                if dEmi_t is None: dEmi_t = root.find(f".//{cte('ide')}/{cte('dEmi')}")
                data_emissao = (dEmi_t.text or "")[:10] if dEmi_t is not None else ""

                emit_t = root.find(f".//{cte('emit')}")
                if emit_t is not None:
                    c = emit_t.find(cte("CNPJ"))
                    cnpj_emitente = c.text if c is not None else "N/A"

                vTPrest_t = root.find(f".//{cte('vPrest')}/{cte('vTPrest')}")
                try:
                    valor = float(vTPrest_t.text) if vTPrest_t is not None and vTPrest_t.text else 0.0
                except:
                    valor = 0.0

                dest_t = root.find(f".//{cte('dest')}")
                if dest_t is not None:
                    cp = dest_t.find(cte("CNPJ"))
                    if cp is None: cp = dest_t.find(cte("CPF"))
                    cnpj_parceiro = cp.text if cp is not None else "N/A"
                    np_t = dest_t.find(cte("xNome"))
                    nome_parceiro = np_t.text if np_t is not None else "N/A"

                itens_lista.append({
                    "nNF": nNF, "nItem": 1,
                    "xProd": "Serviço de Transporte (CT-e)",
                    "NCM": "N/A", "CFOP": "N/A",
                    "qCom": 1, "uCom": "Serv", "vUnCom": valor,
                    "vProd": valor, "CST_ICMS": "N/A", "vICMS": 0.0,
                })

            # ── MDF-e ─────────────────────────────────────────────────────────
            elif infMDFe is not None:
                tipo  = "MDF-e"
                chave = infMDFe.get("Id", "").replace("MDFe", "") or None

                nMDF_t = root.find(f".//{mdf('ide')}/{mdf('nMDF')}")
                nNF    = nMDF_t.text if nMDF_t is not None else "N/A"

                dEmi_t = root.find(f".//{mdf('ide')}/{mdf('dhEmi')}")
                data_emissao = (dEmi_t.text or "")[:10] if dEmi_t is not None else ""

                emit_t = root.find(f".//{mdf('emit')}")
                if emit_t is not None:
                    c = emit_t.find(mdf("CNPJ"))
                    cnpj_emitente = c.text if c is not None else "N/A"
                    np_t = emit_t.find(mdf("xNome"))
                    nome_parceiro = np_t.text if np_t is not None else "N/A"

                valor = 0.0
                itens_lista.append({
                    "nNF": nNF, "nItem": 1,
                    "xProd": "Manifesto de Documentos Fiscais (MDF-e)",
                    "NCM":"N/A","CFOP":"N/A","qCom":1,"uCom":"","vUnCom":0.0,
                    "vProd":0.0,"CST_ICMS":"N/A","vICMS":0.0,
                })

            # ── NFSe (padrão municipal legado + padrão nacional SPED) ──────────
            else:
                # Padrão Nacional SPED: xmlns="http://www.sped.fazenda.gov.br/nfse"
                NFSE_NACIONAL_NS = "http://www.sped.fazenda.gov.br/nfse"
                def nfsen(tag): return f"{{{NFSE_NACIONAL_NS}}}{tag}"

                infnfse_nacional = root.find(f".//{nfsen('infNFSe')}")
                is_nfse_nacional = infnfse_nacional is not None

                # Padrão municipal (vários schemas)
                infnfse = infnfse_nacional or self._find_node_any(root, [
                    ".//{*}InfNfse", ".//{*}infNFSe", ".//{*}InfDeclaracaoPrestacaoServico",
                    ".//{*}Nfse/{*}InfNfse", ".//{*}Nfse"
                ])

                if infnfse is not None:
                    tipo = "NFSE"

                    # ── Padrão Nacional SPED ─────────────────────────────────
                    if is_nfse_nacional:
                        # Chave = Id do infNFSe sem prefixo "NFS"
                        id_attr = infnfse_nacional.get("Id","")
                        chave_nfse = id_attr.replace("NFS","") if id_attr else None

                        nNF_t = infnfse_nacional.find(nfsen("nNFSe"))
                        nNF   = nNF_t.text if nNF_t is not None else "N/A"
                        # Usa chave completa como identificador único
                        chave = chave_nfse if chave_nfse else nNF

                        # Data emissão vem do DPS
                        dEmi_t = root.find(f".//{nfsen('dhEmi')}")
                        data_emissao = (dEmi_t.text or "")[:10] if dEmi_t is not None else ""

                        # cStat: 100=autorizada normal, 101=SUBSTITUTA (nova nota válida),
                        #         102=cancelada, 199=substituída (a nota antiga, raro constar no XML)
                        cStat_t = infnfse_nacional.find(nfsen("cStat"))
                        cStat_val = cStat_t.text if cStat_t is not None else "100"

                        if cStat_val in ("102",):
                            # Cancelada: não conta
                            status = "cancelada"
                            valor  = 0.0
                        elif cStat_val == "101":
                            # cStat=101 = NOTA SUBSTITUTA — ela é a nota VÁLIDA que substituiu
                            # uma nota anterior. Status = autorizada (conta para faturamento)
                            # A nota original (anterior) é que deve ser marcada como substituída
                            status = "autorizada"
                            chSubst_t = root.find(f".//{nfsen('chSubstda')}")
                            if chSubst_t is not None and chSubst_t.text:
                                ch_orig = chSubst_t.text.strip()
                                # Registra o par substituição
                                self.substituicoes[ch_orig] = chave
                                xMotivo_t = root.find(f".//{nfsen('xMotivo')}")
                                motivo = xMotivo_t.text if xMotivo_t is not None else "Substituição"
                                self.audit_log.append({
                                    "tipo": "NFSe SUBSTITUIÇÃO",
                                    "chave": chave,
                                    "chave_ref": ch_orig,
                                    "descricao": f"NFSe nº{nNF} substituiu chave ...{ch_orig[-6:]} | Motivo: {motivo}",
                                    "arquivo": nome_arquivo,
                                    "data": data_emissao,
                                })
                                # Marca nota original como substituída retroativamente
                                if ch_orig in self.documentos_processados:
                                    d_orig = self.documentos_processados[ch_orig]
                                    d_orig["status"] = "substituida"
                                    d_orig["inconsistencia_fiscal"] = f"SUBSTITUÍDA por nNFSe={nNF}"
                                    d_orig["fluxo"] = "SUBSTITUÍDA"
                                    GLOBAL_ITEM_DETAILS.pop(ch_orig, None)
                                else:
                                    # Nota original não processada ainda — registra pendente
                                    # será aplicado quando ela for carregada (ver _pos_processamento)
                                    self.substituicoes["_pendente_" + ch_orig] = (chave, nNF, motivo)
                        else:
                            # cStat=100 = autorizada normal
                            status = "autorizada"

                        # Valor bruto do serviço
                        vServ_t = root.find(f".//{nfsen('vServ')}")
                        vLiq_t  = infnfse_nacional.find(f".//{nfsen('vLiq')}")
                        try:
                            valor = float(vServ_t.text) if vServ_t is not None and vServ_t.text else \
                                    float(vLiq_t.text)  if vLiq_t  is not None and vLiq_t.text  else 0.0
                        except:
                            valor = 0.0

                        # Prestador (emit no padrão nacional)
                        emit_t = infnfse_nacional.find(nfsen("emit"))
                        if emit_t is not None:
                            cnpj_t = emit_t.find(nfsen("CNPJ"))
                            nome_t = emit_t.find(nfsen("xNome"))
                            cnpj_emitente = cnpj_t.text if cnpj_t is not None else "N/A"
                            nome_emit = nome_t.text if nome_t is not None else ""
                        else:
                            nome_emit = ""

                        # Tomador (toma no padrão nacional, dentro do DPS)
                        toma_t = root.find(f".//{nfsen('toma')}")
                        if toma_t is not None:
                            cnpj_t2 = toma_t.find(nfsen("CNPJ"))
                            nome_t2 = toma_t.find(nfsen("xNome"))
                            cnpj_parceiro = cnpj_t2.text if cnpj_t2 is not None else "N/A"
                            nome_parceiro = nome_t2.text if nome_t2 is not None else "N/A"
                        else:
                            nome_parceiro = nome_emit or "N/A"

                        # Serviço
                        cTribNac_t = root.find(f".//{nfsen('cTribNac')}")
                        cNBS_t     = root.find(f".//{nfsen('cNBS')}")
                        xNBS_t     = infnfse_nacional.find(nfsen("xNBS"))
                        item_lista = cTribNac_t.text if cTribNac_t is not None else "N/A"
                        cnae       = "N/A"
                        codtrib    = cNBS_t.text if cNBS_t is not None else "N/A"
                        desc_serv  = xNBS_t.text if xNBS_t is not None else "Serviço Prestado"

                        # Tributação ISS
                        pTotTribSN_t = root.find(f".//{nfsen('pTotTribSN')}")
                        tpRetISSQN_t = root.find(f".//{nfsen('tpRetISSQN')}")
                        try:
                            aliq_iss = float(pTotTribSN_t.text) if pTotTribSN_t is not None else 0.0
                        except:
                            aliq_iss = 0.0
                        ret_iss = valor * (aliq_iss/100) if tpRetISSQN_t is not None and tpRetISSQN_t.text == "1" else 0.0
                        iss_dev = valor * (aliq_iss/100)
                        ret_ir = ret_pis = ret_cofins = ret_csll = ret_inss = 0.0
                        total_ret = ret_iss
                        val_liq = valor - total_ret

                        # Flags fiscais
                        flags = []
                        if ret_iss > 0:
                            flags.append(f"ISS RETIDO R${ret_iss:.2f} - Verificar obrigatoriedade SN")
                        if aliq_iss > 5.0:
                            flags.append(f"Alíq. tributação SN {aliq_iss:.2f}% (conferir apuração)")
                        if cStat_val == "101":
                            flags.append("NFSe SUBSTITUTA (nota válida — substitui nota anterior)")
                        inconsistencia_fiscal = " | ".join(flags) if flags else "OK"

                        ret_str = f" [ISS={ret_iss:.2f} | Líq={val_liq:.2f}]" if total_ret > 0 else ""
                        itens_lista.append({
                            "nNF": nNF, "nItem": 1,
                            "xProd": f"{desc_serv} (cTrib:{item_lista}){ret_str}",
                            "NCM":"N/A","CFOP":"N/A","qCom":1,"uCom":"Serv",
                            "vUnCom": valor, "vProd": valor,
                            "CST_ICMS": f"SN {aliq_iss:.2f}%" if aliq_iss > 0 else "N/A",
                            "vICMS": iss_dev,
                            "nfse_item": item_lista or "N/A",
                            "cnae": cnae,
                            "codtrib": codtrib,
                            "ret_ir": 0., "ret_pis": 0., "ret_cofins": 0.,
                            "ret_csll": 0., "ret_inss": 0., "ret_iss": ret_iss,
                            "valor_liquido": val_liq,
                        })
                        nfse_item_breakdown = (item_lista or "N/A", cnae, codtrib)

                    # ── Padrão Municipal (legado) ─────────────────────────────
                    else:
                        numero = self._find_text_any(infnfse, [
                            ".//{*}Numero", ".//{*}NumeroNfse", ".//{*}NumeroNfSe"
                        ]) or self._find_text_any(root, [".//{*}Numero"])
                        nNF   = numero if numero else "N/A"
                        chave = nNF

                        dEmi_t = self._find_text_any(infnfse, [
                            ".//{*}DataEmissao", ".//{*}Competencia",
                            ".//{*}InfRps/{*}DataEmissao",
                        ])
                        data_emissao = (dEmi_t or "")[:10]

                        if root.find(".//{*}NfseCancelamento") is not None:
                            status = "cancelada"
                            valor  = 0.0
                        else:
                            vbruto_t = self._find_text_any(infnfse, [
                                ".//{*}ValoresNfse/{*}ValorServicos",
                                ".//{*}Valores/{*}ValorServicos",
                                ".//{*}ValorServicos",
                                ".//{*}ValoresNfse/{*}ValorLiquidoNfse",
                            ])
                            try:
                                valor = float(vbruto_t) if vbruto_t else 0.0
                            except:
                                valor = 0.0

                            def _pf(node, paths):
                                t = self._find_text_any(node, paths)
                                try: return float(t) if t else 0.0
                                except: return 0.0

                            ret_ir     = _pf(infnfse, [".//{*}ValoresNfse/{*}ValorIr",".//{*}ValorIr"])
                            ret_pis    = _pf(infnfse, [".//{*}ValoresNfse/{*}ValorPis",".//{*}ValorPis"])
                            ret_cofins = _pf(infnfse, [".//{*}ValoresNfse/{*}ValorCofins",".//{*}ValorCofins"])
                            ret_csll   = _pf(infnfse, [".//{*}ValoresNfse/{*}ValorCsll",".//{*}ValorCsll"])
                            ret_inss   = _pf(infnfse, [".//{*}ValoresNfse/{*}ValorInss",".//{*}ValorInss"])
                            ret_iss    = _pf(infnfse, [".//{*}ValoresNfse/{*}ValorIssRetido",".//{*}ValorIssRetido"])
                            aliq_iss   = _pf(infnfse, [".//{*}ValoresNfse/{*}Aliquota",".//{*}Aliquota"])
                            iss_dev    = _pf(infnfse, [".//{*}ValoresNfse/{*}ValorIss",".//{*}ValorIss"])
                            total_ret  = ret_ir + ret_pis + ret_cofins + ret_csll + ret_inss + ret_iss
                            val_liq    = valor - total_ret

                            flags = []
                            if ret_iss > 0:
                                flags.append(f"ISS RETIDO R${ret_iss:.2f} - Verificar obrigatoriedade SN")
                            if ret_pis > 0 or ret_cofins > 0 or ret_csll > 0:
                                flags.append("RET. PCC INDEVIDA para SN (LC 123/06)")
                            if aliq_iss > 5.0:
                                flags.append(f"Alíq ISS {aliq_iss:.2f}% > 5% (limite legal)")
                            if aliq_iss > 0 and valor > 0:
                                iss_esp = valor * (aliq_iss / 100)
                                if abs(iss_dev - iss_esp) > 0.10:
                                    flags.append(f"ISS apurado R${iss_dev:.2f} diverge de R${iss_esp:.2f}")
                            inconsistencia_fiscal = " | ".join(flags) if flags else "OK"

                            item_lista = self._find_text_any(infnfse, [
                                ".//{*}ItemListaServico",".//{*}Servico/{*}ItemListaServico"
                            ])
                            cnae    = self._find_text_any(infnfse, [".//{*}Servico/{*}CodigoCnae",".//{*}CodigoCnae"])
                            codtrib = self._find_text_any(infnfse, [".//{*}CodigoTributacaoMunicipio",".//{*}CodigoTributacao"])

                            prest_cnpj = self._find_text_any(infnfse, [
                                ".//{*}PrestadorServico/{*}CpfCnpj/{*}Cnpj",
                                ".//{*}Prestador/{*}CpfCnpj/{*}Cnpj",
                            ])
                            prest_nome = self._find_text_any(infnfse, [
                                ".//{*}PrestadorServico/{*}RazaoSocial",
                                ".//{*}Prestador/{*}RazaoSocial",
                            ])
                            tom_cnpj = self._find_text_any(infnfse, [
                                ".//{*}TomadorServico/{*}IdentificacaoTomador/{*}CpfCnpj/{*}Cnpj",
                                ".//{*}TomadorServico/{*}CpfCnpj/{*}Cnpj",
                            ])
                            tom_nome = self._find_text_any(infnfse, [".//{*}TomadorServico/{*}RazaoSocial"])

                            if prest_cnpj: cnpj_emitente = prest_cnpj
                            if tom_cnpj:   cnpj_parceiro = tom_cnpj
                            nome_parceiro = tom_nome or prest_nome or "N/A"

                            ret_str = ""
                            if total_ret > 0:
                                ret_str = (f" [IR={ret_ir:.2f} PIS={ret_pis:.2f}"
                                           f" COF={ret_cofins:.2f} CSLL={ret_csll:.2f}"
                                           f" INSS={ret_inss:.2f} ISS={ret_iss:.2f}"
                                           f" | Líq={val_liq:.2f}]")

                            itens_lista.append({
                                "nNF": nNF, "nItem": 1,
                                "xProd": (f"Serviço: Item {item_lista}" if item_lista else "Serviço Prestado") + ret_str,
                                "NCM":"N/A","CFOP":"N/A","qCom":1,"uCom":"Serv",
                                "vUnCom": valor, "vProd": valor,
                                "CST_ICMS": f"ISS {aliq_iss:.2f}%" if aliq_iss > 0 else "N/A",
                                "vICMS": iss_dev,
                                "nfse_item": item_lista or "N/A",
                                "cnae": cnae or "N/A",
                                "codtrib": codtrib or "N/A",
                                "ret_ir": ret_ir, "ret_pis": ret_pis,
                                "ret_cofins": ret_cofins, "ret_csll": ret_csll,
                                "ret_inss": ret_inss, "ret_iss": ret_iss,
                                "valor_liquido": val_liq,
                            })
                            nfse_item_breakdown = (item_lista or "N/A", cnae or "N/A", codtrib or "N/A")

            # ── Cancelamentos ────────────────────────────────────────────────────────
            # Os XMLs de cancelamento chegam em vários formatos:
            #   A) procEventoNFe: contém <retEvento> com cStat=135/136 e tpEvento=110111
            #   B) eventoNFe puro: contém <evento>/<infEvento>/<tpEvento>=110111
            #   C) nfeProc: contém <protNFe>/<infProt>/<cStat>=101 (nota substituta)
            #   D) XML sem infNFe → tipo="N/A" e pfx=None → cancelamento ignorado (BUG)
            # Solução: detectar pelo namespace do root e buscar em todos os NSs possíveis
            try:
                # ── Detectar tipo pelo namespace do root quando infNFe/infCte ausente ─
                if tipo == "N/A":
                    if NNS in root.tag:
                        tipo = "NF-e"
                    elif CTS in root.tag:
                        tipo = "CT-e"
                    elif MDS in root.tag:
                        tipo = "MDF-e"
                    else:
                        # Último recurso: detectar pelo tpEvento dentro do XML
                        for _ns in (NNS, CTS, MDS):
                            _tpev = root.find(f".//{{{_ns}}}tpEvento")
                            if _tpev is not None:
                                tipo = "NF-e" if _ns == NNS else ("CT-e" if _ns == CTS else "MDF-e")
                                break

                # ── Extrair chave do evento se ainda não temos ───────────────────────
                if chave is None:
                    for _ns in (NNS, CTS, MDS):
                        for _tag in (f"{{{_ns}}}chNFe", f"{{{_ns}}}chCTe", f"{{{_ns}}}chMDFe"):
                            _ch = root.find(f".//{_tag}")
                            if _ch is not None and _ch.text and _ch.text.strip().isdigit():
                                chave = _ch.text.strip()
                                break
                        if chave:
                            break

                # ── Verificação 1: retEvento/infEvento (procEventoNFe/CTe) ──────────
                # cStat 135 = cancelamento homologado, 136 = cancelamento extemporâneo
                if not is_cancellation_event:
                    for _ns in (NNS, CTS, MDS):
                        _inf_ev = root.find(f".//{{{_ns}}}retEvento/{{{_ns}}}infEvento")
                        if _inf_ev is None:
                            _inf_ev = root.find("{*}retEvento/{*}infEvento") or root.find(f".//{{{_ns}}}infEvento[@Id]")
                        if _inf_ev is not None:
                            _cs = _inf_ev.find(f"{{{_ns}}}cStat")
                            if _cs is None: _cs = _inf_ev.find("{*}cStat")
                            _tp = _inf_ev.find(f"{{{_ns}}}tpEvento")
                            if _tp is None: _tp = _inf_ev.find("{*}tpEvento")
                            _csv = (_cs.text or "").strip() if _cs is not None else ""
                            _tpv = (_tp.text or "").strip() if _tp is not None else ""
                            if _csv in ("135", "136") and _tpv == "110111":
                                status = "cancelada"; is_cancellation_event = True
                                break

                # ── Verificação 2: evento/infEvento/tpEvento (eventoNFe puro) ────────
                if not is_cancellation_event:
                    for _ns in (NNS, CTS, MDS):
                        _tp = root.find(f".//{{{_ns}}}evento/{{{_ns}}}infEvento/{{{_ns}}}tpEvento")
                        if _tp is None:
                            _tp = root.find(f".//{{{_ns}}}tpEvento")
                        if _tp is None:
                            _tp = root.find("{*}tpEvento")
                        if _tp is not None and (_tp.text or "").strip() == "110111":
                            status = "cancelada"; is_cancellation_event = True
                            break

                # ── Verificação 3: descEvento contém "Cancelamento" ──────────────────
                if not is_cancellation_event:
                    for _ns in (NNS, CTS, MDS):
                        _desc = root.find(f".//{{{_ns}}}descEvento")
                        if _desc is None:
                            _desc = root.find("{*}descEvento")
                        if _desc is not None and _desc.text:
                            if "CANCELAMENTO" in _desc.text.upper():
                                status = "cancelada"; is_cancellation_event = True
                                break

                # ── Verificação 4: protNFe/infProt/cStat=101 (nota cancelada no envio)
                if not is_cancellation_event:
                    for _ns in (NNS, CTS, MDS):
                        for _prot in (
                            f".//{{{_ns}}}protNFe/{{{_ns}}}infProt/{{{_ns}}}cStat",
                            f".//{{{_ns}}}protCTe/{{{_ns}}}infProt/{{{_ns}}}cStat",
                            f".//{{{_ns}}}protMDFe/{{{_ns}}}infProt/{{{_ns}}}cStat",
                        ):
                            _csp = root.find(_prot)
                            if _csp is not None and (_csp.text or "").strip() == "101":
                                status = "cancelada"; is_cancellation_event = True
                                break
                        if is_cancellation_event:
                            break

                # ── Extrair nNF da chave (dígitos 26-34) se ainda "N/A" ─────────────
                if nNF == "N/A" and chave and len(chave) == 44 and chave.isdigit():
                    try:
                        nNF = str(int(chave[25:34]))
                    except Exception:
                        pass

            except Exception as _ce:
                logging.debug(f"Erro detecção cancelamento {nome_arquivo}: {_ce}")

            # ── Fluxo ─────────────────────────────────────────────────────────
            if cnpj_empresa:
                cei = cnpj_emitente.replace(".","").replace("/","").replace("-","").strip()
                cpi = cnpj_parceiro.replace(".","").replace("/","").replace("-","").strip() if cnpj_parceiro not in ("N/A","") else ""
                if tipo in ("NF-e", "NFC-e"):
                    if cei == cnpj_empresa:
                        if tipo == "NFC-e":
                            fluxo = "Saída Própria"  # NFC-e = sempre venda (tpNF=1)
                        else:
                            fluxo = "Saída Própria" if tpNF == "1" else "Entrada Própria (Dev)"
                    elif cpi == cnpj_empresa:
                        fluxo = "Entrada de Terceiros"
                    else:
                        fluxo = "Terceiros"
                elif tipo in ("CT-e","MDF-e"):
                    if cei == cnpj_empresa:
                        fluxo = f"Saída ({tipo})"
                    elif cpi == cnpj_empresa:
                        fluxo = f"Entrada ({tipo})"
                    else:
                        fluxo = "Terceiros"
                elif tipo == "NFSE":
                    if cei == cnpj_empresa:
                        # Empresa é o prestador → emitiu a nota → SAÍDA (receita)
                        fluxo = "Saída (NFSE)"
                    elif cpi == cnpj_empresa:
                        # Empresa é o tomador → recebeu o serviço → ENTRADA (despesa)
                        fluxo = "Entrada (NFSE)"
                    else:
                        fluxo = "Terceiros"

            if status in ("cancelada",):
                fluxo = "CANCELADO"
            elif status == "substituida":
                fluxo = "SUBSTITUÍDA"

            # ── Forçar cancelamento por nome de arquivo/pasta ─────────────────
            # Quando o usuário importa ZIP/pasta chamado CANCELAD* ou usa o botão
            # "Importar Canceladas", o documento é marcado como cancelado mesmo
            # que o XML original tenha cStat=100 (o evento de cancel é separado)
            if forcar_cancel and status == "autorizada" and not is_cancellation_event:
                status = "cancelada"
                fluxo  = "CANCELADO"
                inconsistencia_fiscal = "CANCELADA"
                is_cancellation_event = True  # evita contabilizar nos totais de faturamento

            # ── Validação de chave ────────────────────────────────────────────
            chave_flag = ""
            if chave and len(chave) == 44 and chave.isdigit():
                if not self._validar_chave_nfe(chave):
                    chave_flag = "CHAVE INVÁLIDA (dígito verificador)"
                    if inconsistencia_fiscal == "OK":
                        inconsistencia_fiscal = chave_flag
                    else:
                        inconsistencia_fiscal += f" | {chave_flag}"

            # ── Armazenamento ─────────────────────────────────────────────────
            novo_doc = {
                "nNF": nNF, "valor": valor, "status": status,
                "tipo": tipo, "arquivo": nome_arquivo,
                "data_emissao": data_emissao,
                "cfop_breakdown": cfop_breakdown,
                "nfse_item_breakdown": nfse_item_breakdown,
                "cnpj_parceiro": cnpj_parceiro,
                "nome_parceiro": nome_parceiro,
                "cnpj_emitente": cnpj_emitente,
                "inconsistencia_fiscal": inconsistencia_fiscal,
                "fluxo": fluxo,
            }

            if is_cancellation_event:
                if chave in self.documentos_processados:
                    d = self.documentos_processados[chave]
                    d["status"] = "cancelada"
                    d["fluxo"] = "CANCELADO"
                    d["inconsistencia_fiscal"] = "CANCELADA"
                    d["arquivo"] += f" / Evento: {nome_arquivo}"
                    GLOBAL_ITEM_DETAILS.pop(chave, None)
                else:
                    novo_doc["inconsistencia_fiscal"] = "CANCELADA"
                    novo_doc["nome_parceiro"] = f"Verificar Chave: {chave}"
                    self.documentos_processados[chave] = novo_doc
            else:
                if chave not in self.documentos_processados or \
                   self.documentos_processados[chave]["status"] != "cancelada":
                    self.documentos_processados[chave] = novo_doc
                    GLOBAL_ITEM_DETAILS[chave] = itens_lista

                if tipo == "NFSE" and novo_doc["status"] == "autorizada" and itens_lista:
                    for item in itens_lista:
                        key = (item.get("nfse_item","N/A"),
                               item.get("cnae","N/A"),
                               item.get("codtrib","N/A"))
                        tot_r = sum(item.get(k, 0.0) for k in
                                    ("ret_ir","ret_pis","ret_cofins","ret_csll","ret_inss","ret_iss"))
                        self.nfse_item_totals[key]["bruto"]    += item.get("vProd", 0.0)
                        self.nfse_item_totals[key]["retencoes"]+= tot_r
                        self.nfse_item_totals[key]["liquido"]  += item.get("valor_liquido",
                                                                            item.get("vProd",0.0)-tot_r)

        except Exception as e:
            logging.exception(f"Erro inesperado {nome_arquivo}: {e}")
            self._registrar_erro(f"Erro inesperado {nome_arquivo}: {e}")

    # ── Validação Chave NF-e (módulo 11) ─────────────────────────────────────────
    def _validar_chave_nfe(self, chave):
        """Valida dígito verificador da chave NF-e/NFC-e (mod 11 SEFAZ).
        Os pesos 2-9 são aplicados nos dígitos lidos da DIREITA para a ESQUERDA."""
        if len(chave) != 44 or not chave.isdigit():
            return False
        pesos = [2,3,4,5,6,7,8,9,2,3,4,5,6,7,8,9,2,3,4,5,6,7,8,9,2,3,4,5,6,7,8,9,2,3,4,5,6,7,8,9,2,3,4]
        soma  = sum(int(c)*p for c,p in zip(reversed(chave[:43]), pesos))
        r     = soma % 11
        dv    = 0 if r < 2 else 11 - r
        return str(dv) == chave[43]

    # ── Quebras de sequência ──────────────────────────────────────────────────────
    def detectar_quebra_sequencia(self):
        # Agrupa por (CNPJ emitente, tipo doc, série) — NF-e e NFC-e têm séries independentes
        seqs = defaultdict(list)
        for chave, doc in self.documentos_processados.items():
            if doc["tipo"] in ("NF-e","NFC-e","NFSE") and doc["status"] not in ("cancelada","substituida"):
                try:
                    nnum = int(str(doc.get("nNF","")).strip())
                    # Extrai série da chave: posições 22-24 (3 dígitos)
                    ch = str(chave)
                    serie = ch[22:25] if len(ch) == 44 and ch.isdigit() else "001"
                    grupo = f"{doc.get('cnpj_emitente','N/A')} | {doc['tipo']} | Série {serie}"
                    seqs[grupo].append(nnum)
                except ValueError:
                    pass
        alerts = []
        for grupo, nums in seqs.items():
            nums_ord = sorted(set(nums))
            dups = [n for n in nums if nums.count(n) > 1]
            for d in sorted(set(dups)):
                alerts.append(f"{grupo}: duplicata nº {d}")
            for i in range(1, len(nums_ord)):
                p, c = nums_ord[i-1], nums_ord[i]
                if c != p + 1:
                    faltam = c - p - 1
                    alerts.append(f"{grupo}: quebra entre {p} e {c} (faltam {faltam})")
        return alerts

    # ── Atualizar Interface ───────────────────────────────────────────────────────
    def atualizar_interface(self):
        self._clear_all_trees()
        self.cfop_totals.clear()

        total_valor  = 0.0   # Faturamento = apenas saídas autorizadas
        counts       = defaultdict(lambda: defaultdict(int))
        partner_acc  = defaultdict(lambda: {"valor": 0.0, "qtd": 0, "ultimo": ""})
        cfop_qtd     = defaultdict(int)
        periodo_acc  = defaultdict(lambda: defaultdict(lambda: {"auth":0,"canc":0,"valor_auth":0.0,"valor_canc":0.0}))
        n_inconsistencias = 0

        for chave, doc in sorted(self.documentos_processados.items(), key=lambda x: (x[0] or "")):
            counts[doc["status"]][doc.get("tipo","N/A")] += 1
            fluxo_doc = doc.get("fluxo","")
            if doc["status"] == "autorizada":
                # Faturamento = saídas próprias. Entradas de serviço NÃO entram no faturamento
                is_entrada = "Entrada" in fluxo_doc
                if not is_entrada:
                    total_valor += doc["valor"]
                for b in doc.get("cfop_breakdown", []):
                    self.cfop_totals[b["cfop"]] += b["vProd"]
                    cfop_qtd[b["cfop"]] += 1
                pk = (doc.get("cnpj_parceiro","N/A"), doc.get("nome_parceiro","N/A"))
                if pk[0] != "N/A":
                    partner_acc[pk]["valor"] += doc["valor"]
                    partner_acc[pk]["qtd"]   += 1
                    d = doc.get("data_emissao","")
                    if d > partner_acc[pk]["ultimo"]:
                        partner_acc[pk]["ultimo"] = d
            inc = doc.get("inconsistencia_fiscal","OK")
            if inc not in ("OK","CANCELADA","") and not inc.startswith("SUBSTITUÍDA"):
                n_inconsistencias += 1

            # Agrupamento por período (MM/AAAA)
            data = doc.get("data_emissao","")
            periodo = data[5:7] + "/" + data[:4] if len(data) >= 7 else "Sem data"
            tp = doc.get("tipo","N/A")
            if doc["status"] == "autorizada":
                periodo_acc[periodo][tp]["auth"]       += 1
                periodo_acc[periodo][tp]["valor_auth"] += doc["valor"]
            else:
                periodo_acc[periodo][tp]["canc"]       += 1
                periodo_acc[periodo][tp]["valor_canc"] += doc["valor"]

        # ── KPIs ────────────────────────────────────────────────────────────
        self._kpi_cards["kpi_total"].update(f"R$ {self._fmt(total_valor)}")
        self._kpi_cards["kpi_nfe"].update(
            f"{counts['autorizada']['NF-e'] + counts['autorizada']['NFC-e']} / "
            f"{counts['cancelada']['NF-e']  + counts['cancelada']['NFC-e']}")
        self._kpi_cards["kpi_cte"].update(
            f"{counts['autorizada']['CT-e']} / {counts['cancelada']['CT-e']}")
        self._kpi_cards["kpi_nfse"].update(
            f"{counts['autorizada']['NFSE']} / {counts['cancelada']['NFSE']}")
        self._kpi_cards["kpi_erros"].update(str(self.erros_detectados))
        self._kpi_cards["kpi_incons"].update(str(n_inconsistencias))

        self.lbl_arquivos.configure(text=f"Arquivos: {self.arquivos_contados}")
        self.lbl_erros_side.configure(text=f"⚠  Erros: {self.erros_detectados}")
        self.lbl_incons_side.configure(text=f"🚩  Inconsistências: {n_inconsistencias}")

        # ── Tabela documentos (com filtros) ──────────────────────────────────
        self.aplicar_filtros()

        # ── Itens ────────────────────────────────────────────────────────────
        alt_i = False
        srch = getattr(self, "item_search_var", None)
        srch_val = srch.get().lower() if srch else ""
        for chave, itens in GLOBAL_ITEM_DETAILS.items():
            doc = self.documentos_processados.get(chave)
            if not doc or doc["status"] != "autorizada":
                continue
            for item in itens:
                xp = item.get("xProd","")
                if srch_val and srch_val not in xp.lower() and srch_val not in item.get("NCM","").lower():
                    continue
                sugestao = item.get("sugestao_trib", "")
                # Determina tag visual
                if "⚠" in sugestao or ("warn" in sugestao.lower() and sugestao):
                    tag = "sugest_alt_warn" if alt_i else "sugest_warn"
                elif sugestao and ("ST" in sugestao or "error" in sugestao):
                    tag = "sugest_alt_error" if alt_i else "sugest_error"
                else:
                    tag = "alt" if alt_i else ()
                alt_i = not alt_i
                self.tree_itens.insert("", "end", tags=(tag,) if tag else (), values=(
                    item.get("nNF","N/A"),
                    xp,
                    item.get("NCM","N/A"),
                    item.get("CEST",""),
                    f"{item.get('qCom',0):.3f}".rstrip("0").rstrip("."),
                    item.get("uCom",""),
                    self._fmt(item.get("vUnCom",0.0)),
                    self._fmt(item.get("vProd",0.0)),
                    item.get("CST_ICMS","N/A"),
                    f"{item.get('CST_PIS','N/A')}/{item.get('CST_COF','N/A')}",
                    self._fmt(item.get("vICMS",0.0)),
                    sugestao,
                ))

        # ── CFOP ─────────────────────────────────────────────────────────────
        alt_c = False
        for cfop, valor in sorted(self.cfop_totals.items()):
            desc = self.CFOP_DESC.get(cfop, "—")
            tags = ("alt",) if alt_c else ()
            alt_c = not alt_c
            self.tree_cfop.insert("", "end", tags=tags, values=(
                cfop, desc, cfop_qtd[cfop], self._fmt(valor)
            ))

        # ── NFSe ─────────────────────────────────────────────────────────────
        alt_n = False
        for key, totais in sorted(self.nfse_item_totals.items()):
            il, cn, ct = key
            tags = ("alt",) if alt_n else ()
            alt_n = not alt_n
            self.tree_nfse.insert("", "end", tags=tags, values=(
                il, cn, ct,
                sum(1 for d in self.documentos_processados.values()
                    if d.get("nfse_item_breakdown") == key and d["status"] == "autorizada"),
                self._fmt(totais["bruto"]),
                self._fmt(totais["retencoes"]),
                self._fmt(totais["liquido"]),
            ))

        # ── Parceiros ────────────────────────────────────────────────────────
        alt_p = False
        for pk, dados in sorted(partner_acc.items(), key=lambda x: -x[1]["valor"]):
            cnpj_p, nome_p = pk
            tags = ("alt",) if alt_p else ()
            alt_p = not alt_p
            self.tree_partner.insert("", "end", tags=tags, values=(
                nome_p, cnpj_p, dados["qtd"],
                self._fmt(dados["valor"]), dados["ultimo"]
            ))

        # ── Por período ───────────────────────────────────────────────────────
        alt_per = False
        for periodo in sorted(periodo_acc.keys()):
            tipos = periodo_acc[periodo]
            tags = ("alt",) if alt_per else ()
            alt_per = not alt_per
            va = sum(v["valor_auth"] for v in tipos.values())
            vc = sum(v["valor_canc"] for v in tipos.values())
            self.tree_periodo.insert("", "end", tags=tags, values=(
                periodo,
                tipos["NF-e"]["auth"],  tipos["NF-e"]["canc"],
                tipos["CT-e"]["auth"],  tipos["CT-e"]["canc"],
                tipos["NFSE"]["auth"],  tipos["NFSE"]["canc"],
                self._fmt(va), self._fmt(vc),
            ))

        # ── Quebras ──────────────────────────────────────────────────────────
        self.quebras_sequencia_alerts = self.detectar_quebra_sequencia()
        self.lbl_quebras_side.configure(
            text=f"🔢  Quebras: {len(self.quebras_sequencia_alerts)}")

        # ── Auditoria ────────────────────────────────────────────────────────
        try:
            self._atualizar_auditoria()
        except Exception:
            pass

    def _clear_all_trees(self):
        for tree in (self.tree, self.tree_itens, self.tree_cfop,
                     self.tree_nfse, self.tree_partner, self.tree_periodo):
            for i in tree.get_children():
                tree.delete(i)
        # Audit trees (may not exist yet on first clear)
        for attr in ("tree_audit_log","tree_conciliacao","tree_substituicoes"):
            t = getattr(self, attr, None)
            if t:
                for i in t.get_children(): t.delete(i)

    # ── Filtros da tabela principal ───────────────────────────────────────────────
    def aplicar_filtros(self):
        for i in self.tree.get_children():
            self.tree.delete(i)

        f     = self.filter_bar.get_filters()
        count = 0
        alt   = False

        for chave, doc in sorted(self.documentos_processados.items(), key=lambda x: (x[0] or "")):
            # Filtros
            if f["tipo"] != "Todos" and doc.get("tipo") != f["tipo"]:
                continue
            if f["status"] != "Todos":
                if f["status"] == "Autorizada"  and doc["status"] != "autorizada":  continue
                if f["status"] == "Cancelada"   and doc["status"] != "cancelada":   continue
                if f["status"] == "Substituída" and doc["status"] != "substituida": continue
            if f["fluxo"] != "Todos" and doc.get("fluxo") != f["fluxo"]:
                continue
            if f["inc_only"] and doc.get("inconsistencia_fiscal","OK") in ("OK","CANCELADA",""):
                continue
            if f["search"]:
                haystack = " ".join(str(v) for v in [
                    chave, doc.get("nNF"), doc.get("nome_parceiro"),
                    doc.get("cnpj_parceiro"), doc.get("arquivo"),
                ]).lower()
                if f["search"] not in haystack:
                    continue
            if f["data_ini"] and doc.get("data_emissao","") < f["data_ini"]:
                continue
            if f["data_fim"] and doc.get("data_emissao","") > f["data_fim"]:
                continue

            # Tag de cor
            st = doc["status"]
            if st == "cancelada":
                tag = "cancelada"
            elif st == "substituida":
                tag = "substituida"
            elif doc.get("inconsistencia_fiscal","OK") not in ("OK","CANCELADA","SUBSTITUÍDA por",""):
                tag = "inconsistente"
            else:
                tag = "alt" if alt else "ok"
                alt = not alt
            count += 1

            status_label = {
                "autorizada":  "Autorizada",
                "cancelada":   "Cancelada",
                "substituida": "Substituída",
            }.get(st, st.title())

            self.tree.insert("", "end", tags=(tag,), values=(
                chave,
                doc.get("tipo","N/A"),
                doc.get("nNF","N/A"),
                doc.get("data_emissao",""),
                self._fmt(doc.get("valor",0.0)),
                status_label,
                doc.get("fluxo","N/A"),
                doc.get("nome_parceiro","N/A"),
                doc.get("cnpj_parceiro","N/A"),
                doc.get("inconsistencia_fiscal","N/A"),
                doc.get("arquivo","N/A"),
            ))

        self.lbl_count_docs.configure(text=f"{count} documento(s) exibido(s)")

    # ── Diagnóstico ───────────────────────────────────────────────────────────────
    def _registrar_erro(self, msg):
        self.erros_detectados += 1
        logging.error(msg)

    def mostrar_quebras_popup(self):
        if not self.quebras_sequencia_alerts:
            messagebox.showinfo("Quebras de Sequência", "Nenhuma quebra detectada.")
            return
        top = tk.Toplevel(self)
        top.title("Quebras de Sequência")
        top.geometry("740x420")
        top.configure(bg=COLORS["bg_card"])

        tk.Label(top, text="Quebras de Sequência Detectadas",
                 bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                 font=FONT_HEADING).pack(padx=16, pady=(16,4), anchor="w")

        txt = tk.Text(top, bg=COLORS["bg_table"], fg=COLORS["text_primary"],
                      font=FONT_MONO, relief="flat", padx=10, pady=10,
                      insertbackground=COLORS["text_primary"])
        txt.pack(fill="both", expand=True, padx=12, pady=8)
        for a in self.quebras_sequencia_alerts:
            txt.insert("end", a + "\n")
        txt.configure(state="disabled")

        StyledButton(top, text="Fechar", style="secondary",
                     command=top.destroy).pack(pady=8)

    # ── Exportar XLSX ─────────────────────────────────────────────────────────────
    def exportar_xlsx(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Exportar XLSX", "openpyxl não está instalado.\nUse: pip install openpyxl")
            return
        if not self.documentos_processados:
            messagebox.showwarning("Exportar XLSX", "Nenhum documento para exportar.")
            return

        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="Salvar relatório XLSX",
        )
        if not caminho:
            return

        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            empresa  = self.nome_empresa_var.get() or "Empresa"
            cnpj_emp = self.cnpj_empresa_var.get()

            # Paleta Excel
            C_HEADER_FILL = PatternFill("solid", fgColor="0D1117")
            C_ACCENT_FILL = PatternFill("solid", fgColor="4F7EFF")
            C_ALT_FILL    = PatternFill("solid", fgColor="161927")
            C_CANCEL_FILL = PatternFill("solid", fgColor="3A1A1A")
            C_WARN_FILL   = PatternFill("solid", fgColor="3A2E00")
            C_OK_FILL     = PatternFill("solid", fgColor="001A14")

            F_WHITE      = Font(name="Segoe UI", color="FFFFFF", bold=True, size=10)
            F_ACCENT     = Font(name="Segoe UI", color="4F7EFF", bold=True, size=11)
            F_BODY       = Font(name="Segoe UI", color="E8EAF6", size=9)
            F_BODY_BOLD  = Font(name="Segoe UI", color="E8EAF6", bold=True, size=9)
            F_RED        = Font(name="Segoe UI", color="FF6B6B", size=9)
            F_YELLOW     = Font(name="Segoe UI", color="FFB347", size=9)
            F_GREEN      = Font(name="Segoe UI", color="00D4AA", size=9)
            F_DIM        = Font(name="Segoe UI", color="8B90C4", size=9)
            F_TITLE      = Font(name="Segoe UI", color="E8EAF6", bold=True, size=13)

            thin = Side(style="thin", color="252840")
            border = Border(bottom=thin)

            def make_header_row(ws, row, cols_labels):
                for c, lbl in enumerate(cols_labels, 1):
                    cell = ws.cell(row=row, column=c, value=lbl)
                    cell.fill   = C_ACCENT_FILL
                    cell.font   = F_WHITE
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = border

            def title_row(ws, text, merge_to=10):
                ws.row_dimensions[1].height = 36
                ws.row_dimensions[2].height = 20
                ws.cell(1,1, value=f"  {empresa}   CNPJ: {cnpj_emp}").font = F_TITLE
                ws.cell(1,1).fill = C_HEADER_FILL
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_to)
                ws.cell(2,1, value=f"  {text}  —  Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = F_DIM
                ws.cell(2,1).fill = C_HEADER_FILL
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=merge_to)

            # ── Aba Resumo ──────────────────────────────────────────────────
            ws_res = wb.create_sheet("Resumo")
            ws_res.sheet_view.showGridLines = False
            ws_res.column_dimensions["A"].width = 40
            ws_res.column_dimensions["B"].width = 25
            title_row(ws_res, "Resumo Geral", 2)

            counts = defaultdict(lambda: defaultdict(int))
            total_val = 0.0
            n_inc = 0
            for doc in self.documentos_processados.values():
                counts[doc["status"]][doc.get("tipo","N/A")] += 1
                if doc["status"] == "autorizada":
                    total_val += doc["valor"]
                if doc.get("inconsistencia_fiscal","OK") not in ("OK","CANCELADA",""):
                    n_inc += 1

            resumo_rows = [
                ("Faturamento Total Autorizado", f"R$ {self._fmt(total_val)}"),
                ("NF-e Autorizadas",  counts["autorizada"]["NF-e"]),
                ("NF-e Canceladas",   counts["cancelada"]["NF-e"]),
                ("CT-e Autorizados",  counts["autorizada"]["CT-e"]),
                ("CT-e Cancelados",   counts["cancelada"]["CT-e"]),
                ("NFSe Autorizadas",  counts["autorizada"]["NFSE"]),
                ("NFSe Canceladas",   counts["cancelada"]["NFSE"]),
                ("Erros",             self.erros_detectados),
                ("Inconsistências",   n_inc),
                ("Quebras de Sequência", len(self.quebras_sequencia_alerts)),
                ("Arquivos Processados", self.arquivos_contados),
            ]
            for r_idx, (k, v) in enumerate(resumo_rows, 4):
                ws_res.cell(r_idx, 1, k).font  = F_BODY_BOLD
                ws_res.cell(r_idx, 2, v).font  = F_ACCENT if r_idx == 4 else F_BODY
                ws_res.cell(r_idx, 1).fill = C_HEADER_FILL
                ws_res.cell(r_idx, 2).fill = C_HEADER_FILL
                ws_res.cell(r_idx, 2).alignment = Alignment(horizontal="right")
                ws_res.row_dimensions[r_idx].height = 20

            # ── Aba Documentos ───────────────────────────────────────────────
            ws_doc = wb.create_sheet("Documentos")
            ws_doc.sheet_view.showGridLines = False
            title_row(ws_doc, "Lista de Documentos", 11)

            headers = ["Chave","Tipo","Nº Doc","Data","Valor (R$)","Status",
                       "Fluxo","Parceiro","CNPJ/CPF","Inconsistência","Arquivo"]
            widths   = [48,  8,  10,  12,  16,  12,  22,  35,  18,  35,  35]
            make_header_row(ws_doc, 3, headers)
            for c, w in enumerate(widths, 1):
                ws_doc.column_dimensions[get_column_letter(c)].width = w

            for r_idx, (chave, doc) in enumerate(sorted(self.documentos_processados.items(), key=lambda x: (x[0] or "")), 4):
                row = [
                    chave,
                    doc.get("tipo"),
                    doc.get("nNF"),
                    doc.get("data_emissao",""),
                    doc.get("valor",0.0),
                    doc["status"].title(),
                    doc.get("fluxo"),
                    doc.get("nome_parceiro"),
                    doc.get("cnpj_parceiro"),
                    doc.get("inconsistencia_fiscal","OK"),
                    doc.get("arquivo"),
                ]
                inc = doc.get("inconsistencia_fiscal","OK")
                is_cancel = doc["status"] == "cancelada"

                for c_idx, val in enumerate(row, 1):
                    cell = ws_doc.cell(r_idx, c_idx, val)
                    cell.fill = C_CANCEL_FILL if is_cancel else (
                                C_WARN_FILL   if inc not in ("OK","CANCELADA","") else
                                C_ALT_FILL    if r_idx % 2 == 0 else C_HEADER_FILL)
                    cell.font = F_RED    if is_cancel else (
                                F_YELLOW if inc not in ("OK","CANCELADA","") else F_BODY)
                    cell.alignment = Alignment(vertical="center")
                    if c_idx == 5:  # Valor
                        cell.number_format = "#,##0.00"
                ws_doc.row_dimensions[r_idx].height = 18

            ws_doc.freeze_panes = "A4"
            ws_doc.auto_filter.ref = f"A3:K3"

            # ── Aba Itens ────────────────────────────────────────────────────
            ws_it = wb.create_sheet("Itens_Produtos")
            ws_it.sheet_view.showGridLines = False
            title_row(ws_it, "Itens e Produtos (Docs Autorizados)", 10)
            h_it = ["Nº Doc","Produto / Serviço","NCM","Qtd","Un",
                    "Valor Unit. (R$)","Valor Total (R$)","CST/CSOSN","vICMS (R$)","Sugestão Tributação"]
            w_it = [10, 45, 12, 10, 6, 16, 16, 10, 14, 55]
            make_header_row(ws_it, 3, h_it)
            for c, w in enumerate(w_it, 1):
                ws_it.column_dimensions[get_column_letter(c)].width = w

            r_it = 4
            for chave, itens in GLOBAL_ITEM_DETAILS.items():
                doc = self.documentos_processados.get(chave)
                if not doc or doc["status"] != "autorizada":
                    continue
                for item in itens:
                    sugestao = item.get("sugestao_trib", "")
                    row = [
                        item.get("nNF"),
                        item.get("xProd"),
                        item.get("NCM"),
                        item.get("qCom",0),
                        item.get("uCom",""),
                        item.get("vUnCom",0.0),
                        item.get("vProd",0.0),
                        item.get("CST_ICMS"),
                        item.get("vICMS",0.0),
                        sugestao,
                    ]
                    has_sugest = bool(sugestao)
                    for c_idx, val in enumerate(row, 1):
                        cell = ws_it.cell(r_it, c_idx, val)
                        cell.fill = C_WARN_FILL if has_sugest else (C_ALT_FILL if r_it % 2 == 0 else C_HEADER_FILL)
                        cell.font = F_YELLOW if has_sugest else F_BODY
                        cell.alignment = Alignment(vertical="center")
                        if c_idx in (6, 7, 9):
                            cell.number_format = "#,##0.00"
                    ws_it.row_dimensions[r_it].height = 18
                    r_it += 1

            ws_it.freeze_panes = "A4"

            # ── Aba CFOP ─────────────────────────────────────────────────────
            ws_cf = wb.create_sheet("CFOP")
            ws_cf.sheet_view.showGridLines = False
            title_row(ws_cf, "Totalização por CFOP (NF-e Autorizadas)", 4)
            make_header_row(ws_cf, 3, ["CFOP","Descrição","Qtd Docs","Valor Total (R$)"])
            ws_cf.column_dimensions["A"].width = 10
            ws_cf.column_dimensions["B"].width = 60
            ws_cf.column_dimensions["C"].width = 12
            ws_cf.column_dimensions["D"].width = 20

            cfop_qtd_local = defaultdict(int)
            for doc in self.documentos_processados.values():
                if doc["status"] == "autorizada":
                    for b in doc.get("cfop_breakdown",[]):
                        cfop_qtd_local[b["cfop"]] += 1

            for r_cf, (cfop, valor) in enumerate(sorted(self.cfop_totals.items()), 4):
                row = [cfop, self.CFOP_DESC.get(cfop,"—"), cfop_qtd_local[cfop], valor]
                for c_idx, val in enumerate(row, 1):
                    cell = ws_cf.cell(r_cf, c_idx, val)
                    cell.fill = C_ALT_FILL if r_cf % 2 == 0 else C_HEADER_FILL
                    cell.font = F_BODY
                    if c_idx == 4:
                        cell.number_format = "#,##0.00"
                        cell.font = F_ACCENT
                ws_cf.row_dimensions[r_cf].height = 18

            # ── Aba NFSe ─────────────────────────────────────────────────────
            ws_nf = wb.create_sheet("NFSe")
            ws_nf.sheet_view.showGridLines = False
            title_row(ws_nf, "NFSe por Item/CNAE/Código", 6)
            make_header_row(ws_nf, 3, ["Item Lista","CNAE","Cód. Trib.","Bruto (R$)","Retenções (R$)","Líquido (R$)"])
            for c, w in enumerate([18,14,18,18,18,18], 1):
                ws_nf.column_dimensions[get_column_letter(c)].width = w

            for r_nf, (key, totais) in enumerate(sorted(self.nfse_item_totals.items()), 4):
                row = [key[0], key[1], key[2],
                       totais["bruto"], totais["retencoes"], totais["liquido"]]
                for c_idx, val in enumerate(row, 1):
                    cell = ws_nf.cell(r_nf, c_idx, val)
                    cell.fill = C_ALT_FILL if r_nf % 2 == 0 else C_HEADER_FILL
                    cell.font = F_BODY
                    if c_idx in (4,5,6):
                        cell.number_format = "#,##0.00"
                ws_nf.row_dimensions[r_nf].height = 18

            # ── Aba Inconsistências ──────────────────────────────────────────
            ws_inc = wb.create_sheet("Inconsistências")
            ws_inc.sheet_view.showGridLines = False
            title_row(ws_inc, "Documentos com Inconsistências Fiscais", 6)
            make_header_row(ws_inc, 3, ["Chave","Tipo","Nº Doc","Data","Valor (R$)","Inconsistência"])
            for c, w in enumerate([48,8,10,12,16,55], 1):
                ws_inc.column_dimensions[get_column_letter(c)].width = w

            r_inc = 4
            for chave, doc in sorted(self.documentos_processados.items(), key=lambda x: (x[0] or "")):
                inc = doc.get("inconsistencia_fiscal","OK")
                if inc in ("OK","CANCELADA",""):
                    continue
                row = [chave, doc.get("tipo"), doc.get("nNF"),
                       doc.get("data_emissao",""), doc.get("valor",0.0), inc]
                for c_idx, val in enumerate(row, 1):
                    cell = ws_inc.cell(r_inc, c_idx, val)
                    cell.fill = C_WARN_FILL
                    cell.font = F_YELLOW
                    if c_idx == 5:
                        cell.number_format = "#,##0.00"
                ws_inc.row_dimensions[r_inc].height = 18
                r_inc += 1

            # ── Aba Quebras ──────────────────────────────────────────────────
            ws_qb = wb.create_sheet("Quebras_Sequência")
            ws_qb.sheet_view.showGridLines = False
            title_row(ws_qb, "Quebras de Sequência Detectadas", 2)
            ws_qb.column_dimensions["A"].width = 80

            for r_qb, alerta in enumerate(self.quebras_sequencia_alerts, 4):
                cell = ws_qb.cell(r_qb, 1, alerta)
                cell.fill = C_WARN_FILL
                cell.font = F_YELLOW
                ws_qb.row_dimensions[r_qb].height = 20

            # Ajuste geral
            for ws in wb.worksheets:
                ws.sheet_view.showRowColHeaders = False

            wb.save(caminho)
            messagebox.showinfo("Exportar XLSX", f"Relatório salvo:\n{caminho}")

        except Exception as e:
            logging.exception(f"Erro ao exportar XLSX: {e}")
            messagebox.showerror("Exportar XLSX", f"Erro ao gerar XLSX:\n{e}")

    # ── Exportar CSV ──────────────────────────────────────────────────────────────
    def exportar_csv(self):
        if not self.documentos_processados:
            messagebox.showwarning("CSV", "Nenhum documento para exportar.")
            return
        caminho = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            title="Salvar CSV",
        )
        if not caminho:
            return
        try:
            with open(caminho, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(["Chave","Tipo","Nº Doc","Data","Valor","Status","Fluxo",
                             "Parceiro","CNPJ/CPF","Inconsistência","Arquivo"])
                for chave, doc in sorted(self.documentos_processados.items(), key=lambda x: (x[0] or "")):
                    w.writerow([
                        chave, doc.get("tipo"), doc.get("nNF"),
                        doc.get("data_emissao",""),
                        f"{doc.get('valor',0.0):.2f}",
                        doc.get("status"), doc.get("fluxo"),
                        doc.get("nome_parceiro"), doc.get("cnpj_parceiro"),
                        doc.get("inconsistencia_fiscal"), doc.get("arquivo"),
                    ])
            messagebox.showinfo("CSV", f"CSV salvo:\n{caminho}")
        except Exception as e:
            logging.exception(f"Erro CSV: {e}")
            messagebox.showerror("CSV", f"Erro: {e}")

    # ── Exportar PDF ──────────────────────────────────────────────────────────────
    def exportar_pdf(self):
        if not REPORTLAB_AVAILABLE:
            messagebox.showerror("PDF", "reportlab não instalado.\nUse: pip install reportlab")
            return
        if not self.documentos_processados:
            messagebox.showwarning("PDF", "Nenhum documento para exportar.")
            return
        caminho = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            title="Salvar Relatório PDF",
        )
        if not caminho:
            return
        try:
            doc_pdf = SimpleDocTemplate(
                caminho, pagesize=landscape(A4),
                leftMargin=1.5*cm, rightMargin=1.5*cm,
                topMargin=2*cm, bottomMargin=1.5*cm,
            )
            styles = getSampleStyleSheet()
            story  = []

            titulo_style = ParagraphStyle("Titulo", parent=styles["Normal"],
                fontSize=16, fontName="Helvetica-Bold",
                textColor=colors.HexColor("#4F7EFF"), spaceAfter=4)
            sub_style = ParagraphStyle("Sub", parent=styles["Normal"],
                fontSize=9, fontName="Helvetica",
                textColor=colors.HexColor("#8B90C4"), spaceAfter=14)
            h2_style = ParagraphStyle("H2", parent=styles["Normal"],
                fontSize=11, fontName="Helvetica-Bold",
                textColor=colors.HexColor("#E8EAF6"), spaceBefore=12, spaceAfter=6)
            body_style = ParagraphStyle("Body", parent=styles["Normal"],
                fontSize=8, fontName="Helvetica",
                textColor=colors.HexColor("#E8EAF6"))

            empresa = self.nome_empresa_var.get() or "Empresa"
            cnpj_e  = self.cnpj_empresa_var.get()
            agora   = datetime.now().strftime("%d/%m/%Y %H:%M")

            story.append(Paragraph(f"Relatório Fiscal — {empresa}", titulo_style))
            story.append(Paragraph(f"CNPJ: {cnpj_e}   |   Gerado em: {agora}", sub_style))
            story.append(HRFlowable(width="100%", thickness=1,
                                     color=colors.HexColor("#4F7EFF")))
            story.append(Spacer(1, 12))

            # KPI resumo
            counts = defaultdict(lambda: defaultdict(int))
            total_val = 0.0
            n_inc = 0
            for doc in self.documentos_processados.values():
                counts[doc["status"]][doc.get("tipo","N/A")] += 1
                if doc["status"] == "autorizada":
                    total_val += doc["valor"]
                if doc.get("inconsistencia_fiscal","OK") not in ("OK","CANCELADA",""):
                    n_inc += 1

            kpi_data = [
                ["Faturamento Total Auth.", f"R$ {self._fmt(total_val)}"],
                ["NF-e Auth/Canc",          f"{counts['autorizada']['NF-e']} / {counts['cancelada']['NF-e']}"],
                ["CT-e Auth/Canc",           f"{counts['autorizada']['CT-e']} / {counts['cancelada']['CT-e']}"],
                ["NFSe Auth/Canc",           f"{counts['autorizada']['NFSE']} / {counts['cancelada']['NFSE']}"],
                ["Inconsistências",          str(n_inc)],
                ["Arquivos Processados",     str(self.arquivos_contados)],
            ]
            kpi_table = Table(kpi_data, colWidths=[7*cm, 5*cm])
            kpi_table.setStyle(TableStyle([
                ("BACKGROUND",  (0,0), (-1,-1), colors.HexColor("#1A1D2E")),
                ("TEXTCOLOR",   (0,0), (0,-1),  colors.HexColor("#8B90C4")),
                ("TEXTCOLOR",   (1,0), (1,-1),  colors.HexColor("#4F7EFF")),
                ("FONTNAME",    (0,0), (-1,-1),  "Helvetica"),
                ("FONTSIZE",    (0,0), (-1,-1),  9),
                ("ROWBACKGROUNDS",(0,0),(-1,-1),[colors.HexColor("#1A1D2E"),colors.HexColor("#161927")]),
                ("GRID",        (0,0), (-1,-1),  0.5, colors.HexColor("#252840")),
                ("LEFTPADDING", (0,0), (-1,-1),  8),
                ("RIGHTPADDING",(0,0), (-1,-1),  8),
                ("TOPPADDING",  (0,0), (-1,-1),  5),
                ("BOTTOMPADDING",(0,0),(-1,-1),  5),
            ]))
            story.append(kpi_table)
            story.append(Spacer(1, 18))

            # Tabela documentos
            story.append(Paragraph("Documentos Processados", h2_style))
            doc_headers = ["Chave","Tipo","Nº","Data","Valor (R$)","Status","Parceiro","Inconsistência"]
            doc_rows = [doc_headers]
            for chave, doc in sorted(self.documentos_processados.items(), key=lambda x: (x[0] or "")):
                doc_rows.append([
                    chave[-20:] + "..." if len(str(chave)) > 20 else str(chave),
                    doc.get("tipo",""),
                    doc.get("nNF",""),
                    doc.get("data_emissao",""),
                    f"R$ {self._fmt(doc.get('valor',0.0))}",
                    doc["status"].title(),
                    (doc.get("nome_parceiro","") or "")[:30],
                    (doc.get("inconsistencia_fiscal","") or "")[:30],
                ])
            col_w = [5*cm,1.5*cm,1.5*cm,2.2*cm,3*cm,2*cm,4.5*cm,5*cm]
            t = Table(doc_rows, colWidths=col_w, repeatRows=1)
            ts = TableStyle([
                ("BACKGROUND",  (0,0), (-1,0),  colors.HexColor("#4F7EFF")),
                ("TEXTCOLOR",   (0,0), (-1,0),  colors.white),
                ("FONTNAME",    (0,0), (-1,0),  "Helvetica-Bold"),
                ("FONTSIZE",    (0,0), (-1,-1),  7),
                ("FONTNAME",    (0,1), (-1,-1),  "Helvetica"),
                ("TEXTCOLOR",   (0,1), (-1,-1),  colors.HexColor("#E8EAF6")),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.HexColor("#1A1D2E"),colors.HexColor("#161927")]),
                ("GRID",        (0,0), (-1,-1),  0.4, colors.HexColor("#252840")),
                ("TOPPADDING",  (0,0), (-1,-1),  4),
                ("BOTTOMPADDING",(0,0),(-1,-1),  4),
                ("LEFTPADDING", (0,0), (-1,-1),  4),
            ])
            # Colorir canceladas e inconsistentes
            for r_idx in range(1, len(doc_rows)):
                row_data = doc_rows[r_idx]
                if row_data[5] == "Cancelada":
                    ts.add("TEXTCOLOR", (0,r_idx),(-1,r_idx), colors.HexColor("#FF6B6B"))
                elif row_data[7] not in ("OK","CANCELADA",""):
                    ts.add("TEXTCOLOR", (0,r_idx),(-1,r_idx), colors.HexColor("#FFB347"))
            t.setStyle(ts)
            story.append(t)

            # Rodapé
            story.append(Spacer(1, 16))
            story.append(HRFlowable(width="100%", thickness=0.5,
                                     color=colors.HexColor("#252840")))
            story.append(Paragraph(
                f"FiscalPro Analytics v{APP_VERSION}  —  {empresa}  —  {agora}",
                ParagraphStyle("Footer", parent=styles["Normal"], fontSize=7,
                               fontName="Helvetica", textColor=colors.HexColor("#4A4F7A"),
                               alignment=TA_CENTER)
            ))

            doc_pdf.build(story)
            messagebox.showinfo("PDF", f"Relatório PDF salvo:\n{caminho}")

        except Exception as e:
            logging.exception(f"Erro PDF: {e}")
            messagebox.showerror("PDF", f"Erro ao gerar PDF:\n{e}")

    # ── Sessão ────────────────────────────────────────────────────────────────────
    def salvar_sessao(self):
        if not self.documentos_processados:
            messagebox.showwarning("Sessão", "Nada para salvar.")
            return
        caminho = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("Sessão FiscalPro","*.json")],
            title="Salvar Sessão",
        )
        if not caminho:
            return
        try:
            sessao = {
                "versao": APP_VERSION,
                "data_salvo": datetime.now().isoformat(),
                "empresa": self.nome_empresa_var.get(),
                "cnpj": self.cnpj_empresa_var.get(),
                "documentos": {k: v for k,v in self.documentos_processados.items()},
                "nfse_totals": {str(k): v for k,v in self.nfse_item_totals.items()},
                "arquivos_contados": self.arquivos_contados,
                "erros": self.erros_detectados,
            }
            with open(caminho, "w", encoding="utf-8") as f:
                json.dump(sessao, f, ensure_ascii=False, indent=2, default=str)
            messagebox.showinfo("Sessão", f"Sessão salva:\n{caminho}")
        except Exception as e:
            messagebox.showerror("Sessão", f"Erro ao salvar: {e}")

    def carregar_sessao(self):
        caminho = filedialog.askopenfilename(
            filetypes=[("Sessão FiscalPro","*.json"),("Todos","*.*")],
            title="Carregar Sessão",
        )
        if not caminho:
            return
        try:
            with open(caminho, "r", encoding="utf-8") as f:
                sessao = json.load(f)
            self.limpar_resultados()
            self.documentos_processados = sessao.get("documentos", {})
            self.nome_empresa_var.set(sessao.get("empresa",""))
            self.cnpj_empresa_var.set(sessao.get("cnpj",""))
            self.arquivos_contados = sessao.get("arquivos_contados", 0)
            self.erros_detectados  = sessao.get("erros", 0)
            self.atualizar_interface()
            messagebox.showinfo("Sessão", f"Sessão carregada:\n{caminho}")
        except Exception as e:
            messagebox.showerror("Sessão", f"Erro ao carregar: {e}")

    # ── Limpar ────────────────────────────────────────────────────────────────────
    def limpar_resultados(self, clear_files=True):
        self.documentos_processados.clear()
        self.cfop_totals.clear()
        self.nfse_item_totals = defaultdict(lambda: {"bruto":0.,"retencoes":0.,"liquido":0.})
        self.partner_totals.clear()
        GLOBAL_ITEM_DETAILS.clear()
        self.quebras_sequencia_alerts.clear()
        self.substituicoes.clear()
        self.audit_log.clear()
        if clear_files:
            self.arquivos_contados = 0
            self.erros_detectados  = 0
        try:
            self.atualizar_interface()
        except Exception:
            pass

    # ── Config Persistente ────────────────────────────────────────────────────────
    def _load_config(self):
        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception:
            pass
        return {}

    def _salvar_config(self):
        try:
            cfg = {
                "cnpj":        self.cnpj_empresa_var.get(),
                "nome_empresa":self.nome_empresa_var.get(),
                "pasta_padrao":self.pasta_padrao,
            }
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def on_close(self):
        self._salvar_config()
        self.destroy()


# ─── Entry Point ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = App()
    app.protocol("WM_DELETE_WINDOW", app.on_close)
    app.mainloop()
